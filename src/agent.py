import asyncio
import inspect
import logging
import math
import re
import textwrap
import time
from datetime import datetime, timedelta
from pathlib import Path

from dateutil import parser as dateutil_parser
from dotenv import load_dotenv
from livekit import api
from livekit.agents import (
    Agent,
    AgentServer,
    AgentSession,
    JobContext,
    JobProcess,
    RunContext,
    StopResponse,
    cli,
    function_tool,
    get_job_context,
    inference,
    room_io,
    stt,
)
from livekit.agents.inference import TurnDetector
from livekit.agents.llm import ChatContext
from livekit.plugins import ai_coustics, sarvam, silero
from livekit.plugins.turn_detector.multilingual import MultilingualModel
from openpyxl import load_workbook

logger = logging.getLogger("agent")

load_dotenv(".env.local")

# Excel source for customer details, resolved relative to the project root so it
# works regardless of the current working directory.
CUSTOMER_DATA_FILE = Path(__file__).parent.parent / "data" / "Customers.xlsx"

# Name of the customer loaded for this call. Used to locate the row to update
# when persisting a Promise To Pay. We assume one customer is loaded per call.
LOADED_CUSTOMER_NAME = ""


def _is_missing(value) -> bool:
    """Return True for None, NaN, or blank cells."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


# Hindi words for 0-99. Indian numbers in this range are irregular (each has a
# unique word), so they are listed explicitly. Used to SPEAK amounts and dates
# fully in Hindi (Devanagari) — never half-English like "18 hazaar 750".
_HINDI_NUMBERS_0_99 = {
    0: "शून्य",
    1: "एक",
    2: "दो",
    3: "तीन",
    4: "चार",
    5: "पाँच",
    6: "छह",
    7: "सात",
    8: "आठ",
    9: "नौ",
    10: "दस",
    11: "ग्यारह",
    12: "बारह",
    13: "तेरह",
    14: "चौदह",
    15: "पंद्रह",
    16: "सोलह",
    17: "सत्रह",
    18: "अठारह",
    19: "उन्नीस",
    20: "बीस",
    21: "इक्कीस",
    22: "बाईस",
    23: "तेईस",
    24: "चौबीस",
    25: "पच्चीस",
    26: "छब्बीस",
    27: "सत्ताईस",
    28: "अट्ठाईस",
    29: "उनतीस",
    30: "तीस",
    31: "इकतीस",
    32: "बत्तीस",
    33: "तैंतीस",
    34: "चौंतीस",
    35: "पैंतीस",
    36: "छत्तीस",
    37: "सैंतीस",
    38: "अड़तीस",
    39: "उनतालीस",
    40: "चालीस",
    41: "इकतालीस",
    42: "बयालीस",
    43: "तैंतालीस",
    44: "चौवालीस",
    45: "पैंतालीस",
    46: "छियालीस",
    47: "सैंतालीस",
    48: "अड़तालीस",
    49: "उनचास",
    50: "पचास",
    51: "इक्यावन",
    52: "बावन",
    53: "तिरपन",
    54: "चौवन",
    55: "पचपन",
    56: "छप्पन",
    57: "सत्तावन",
    58: "अट्ठावन",
    59: "उनसठ",
    60: "साठ",
    61: "इकसठ",
    62: "बासठ",
    63: "तिरसठ",
    64: "चौंसठ",
    65: "पैंसठ",
    66: "छियासठ",
    67: "सड़सठ",
    68: "अड़सठ",
    69: "उनहत्तर",
    70: "सत्तर",
    71: "इकहत्तर",
    72: "बहत्तर",
    73: "तिहत्तर",
    74: "चौहत्तर",
    75: "पचहत्तर",
    76: "छिहत्तर",
    77: "सतहत्तर",
    78: "अठहत्तर",
    79: "उन्यासी",
    80: "अस्सी",
    81: "इक्यासी",
    82: "बयासी",
    83: "तिरासी",
    84: "चौरासी",
    85: "पचासी",
    86: "छियासी",
    87: "सत्तासी",
    88: "अट्ठासी",
    89: "नवासी",
    90: "नब्बे",
    91: "इक्यानवे",
    92: "बानवे",
    93: "तिरानवे",
    94: "चौरानवे",
    95: "पंचानवे",
    96: "छियानवे",
    97: "सत्तानवे",
    98: "अट्ठानवे",
    99: "निन्यानवे",
}


def _number_to_hindi_words(amount: int) -> str:
    """Spell an integer in Indian-system Hindi words (18750 -> 'अठारह हज़ार सात सौ पचास')."""
    if amount < 0:
        return "माइनस " + _number_to_hindi_words(-amount)
    if amount == 0:
        return _HINDI_NUMBERS_0_99[0]

    crore, rem = divmod(amount, 10_000_000)
    lakh, rem = divmod(rem, 100_000)
    thousand, rem = divmod(rem, 1_000)
    hundred, rem = divmod(rem, 100)  # rem is now 0-99

    parts: list[str] = []
    if crore:
        # crore can exceed 99 for very large values; recurse so it still reads.
        parts += [_number_to_hindi_words(crore), "करोड़"]
    if lakh:
        parts += [_HINDI_NUMBERS_0_99[lakh], "लाख"]
    if thousand:
        parts += [_HINDI_NUMBERS_0_99[thousand], "हज़ार"]
    if hundred:
        parts += [_HINDI_NUMBERS_0_99[hundred], "सौ"]
    if rem:
        parts.append(_HINDI_NUMBERS_0_99[rem])
    return " ".join(parts)


def _format_indian_amount(value) -> str:
    """Format a rupee amount fully in spoken Hindi (18750 -> 'अठारह हज़ार सात सौ पचास रुपये')."""
    if _is_missing(value):
        return "N/A"
    try:
        amount = int(float(value))
    except (TypeError, ValueError):
        return str(value).strip()
    return f"{_number_to_hindi_words(amount)} रुपये"


def _format_date(value) -> str:
    """Format a due date as e.g. '15 June'."""
    if _is_missing(value):
        return "N/A"
    if isinstance(value, datetime):
        return f"{value.day} {value.strftime('%B')}"
    try:
        dt = dateutil_parser.parse(str(value), dayfirst=True, fuzzy=True)
        return f"{dt.day} {dt.strftime('%B')}"
    except (ValueError, TypeError, OverflowError):
        return str(value).strip()


def _load_customer_context(path: Path) -> str:
    """Build CUSTOMER_CONTEXT from the second data row of the Excel sheet.

    Single streaming pass with openpyxl (read_only) instead of two full
    pandas parses -- keeps worker startup latency to a minimum.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    header_found = False
    columns: dict[str, int] = {}
    target_values: tuple | None = None
    data_rows_seen = 0

    for sheet_row in worksheet.iter_rows(values_only=True):
        if not header_found:
            labels = [str(v).strip().lower() for v in sheet_row if v is not None]
            if "customer name" in labels:
                header_found = True
                columns = {
                    str(v).strip().lower().replace(" ", "_"): idx
                    for idx, v in enumerate(sheet_row)
                    if v is not None
                }
            continue
        data_rows_seen += 1
        if data_rows_seen == 2:  # matches the previous df.iloc[1]
            target_values = sheet_row
            break

    workbook.close()

    if not header_found or target_values is None:
        raise RuntimeError(f"Could not locate customer row in {path}")

    def cell(col_name: str):
        idx = columns.get(col_name)
        return (
            target_values[idx] if idx is not None and idx < len(target_values) else None
        )

    name_parts = [
        str(cell(col)).strip()
        for col in ("customer_name", "last_name")
        if not _is_missing(cell(col))
    ]
    customer_name = " ".join(name_parts) or "Customer"

    # Remember the raw name so the PTP writer can find this exact row later.
    global LOADED_CUSTOMER_NAME
    LOADED_CUSTOMER_NAME = customer_name

    loan_amount = _format_indian_amount(cell("loan_amount"))
    due_amount = _format_indian_amount(cell("due_amount"))
    due_date = _format_date(cell("due_date"))

    # Lender is fully data-driven: whatever the sheet says, with a neutral
    # fallback so the prompt never references a hardcoded bank.
    lender_value = cell("lender_name")
    lender_name = "N/A" if _is_missing(lender_value) else str(lender_value).strip()

    return textwrap.dedent(f"""
        Customer Name: {customer_name}
        Lender Name: {lender_name}
        Loan Amount: {loan_amount}
        Outstanding Amount: {due_amount}
        EMI Due Date: {due_date}
    """).strip()


CUSTOMER_CONTEXT = _load_customer_context(CUSTOMER_DATA_FILE)


# ---------------------------------------------------------------------------
# PTP normalization
#
# Excel must always store structured English data: PTP Amount as a plain number
# string ("5000") and PTP Date as DD-MM-YYYY ("25-06-2026"). The spoken
# conversation stays Hindi/Hinglish; only the values handed to
# record_promise_to_pay() and written to Excel are normalized here.
#
# Date policy: we always emit DD-MM-YYYY. When the source value has no year
# (e.g. "25 June", "पच्चीस जून"), we assume the current year. The LLM is
# instructed to resolve relative dates ("next Monday") to an absolute date with
# a year before calling the tool.
# ---------------------------------------------------------------------------

# Devanagari digits -> ASCII digits (२५ -> 25).
_DEV_DIGITS = str.maketrans("०१२३४५६७८९", "0123456789")

# Hindi number words. Includes the common 0-32 (covers day-of-month) plus the
# round tens, which is enough for the amounts and dates seen in this workflow.
_HINDI_UNITS = {
    "शून्य": 0,
    "एक": 1,
    "दो": 2,
    "तीन": 3,
    "चार": 4,
    "पाँच": 5,
    "पांच": 5,
    "छह": 6,
    "छः": 6,
    "छे": 6,
    "सात": 7,
    "आठ": 8,
    "नौ": 9,
    "दस": 10,
    "ग्यारह": 11,
    "बारह": 12,
    "तेरह": 13,
    "चौदह": 14,
    "पंद्रह": 15,
    "पन्द्रह": 15,
    "सोलह": 16,
    "सत्रह": 17,
    "अठारह": 18,
    "उन्नीस": 19,
    "बीस": 20,
    "इक्कीस": 21,
    "बाईस": 22,
    "तेईस": 23,
    "चौबीस": 24,
    "पच्चीस": 25,
    "पचीस": 25,
    "छब्बीस": 26,
    "सत्ताईस": 27,
    "अट्ठाईस": 28,
    "उनतीस": 29,
    "तीस": 30,
    "इकतीस": 31,
    "बत्तीस": 32,
    "चालीस": 40,
    "पचास": 50,
    "साठ": 60,
    "सत्तर": 70,
    "अस्सी": 80,
    "नब्बे": 90,
}
_HINDI_HUNDRED = {"सौ": 100}
_HINDI_BIG = {
    "हज़ार": 1_000,
    "हजार": 1_000,
    "लाख": 100_000,
    "करोड़": 10_000_000,
    "करोड": 10_000_000,
}
# Currency / filler words to ignore while parsing numbers.
_NUMBER_NOISE = {
    "रुपये",
    "रुपए",
    "रुपया",
    "रुपय",
    "रूपये",
    "रूपए",
    "रु",
    "रु.",
    "को",
    "और",
    "मात्र",
    "only",
    "rupaye",
    "rupees",
    "rs",
    "rs.",
}
_HINDI_MONTHS = {
    "जनवरी": "January",
    "फरवरी": "February",
    "फ़रवरी": "February",
    "मार्च": "March",
    "अप्रैल": "April",
    "मई": "May",
    "जून": "June",
    "जुलाई": "July",
    "अगस्त": "August",
    "सितंबर": "September",
    "सितम्बर": "September",
    "अक्टूबर": "October",
    "नवंबर": "November",
    "नवम्बर": "November",
    "दिसंबर": "December",
    "दिसम्बर": "December",
}

# Canonical Hindi word for each day-of-month (1-31), used to SPEAK a date aloud
# ("परसों" -> "एक जुलाई"). Reuses the 0-99 number words for a single source of truth.
_HINDI_DAY_WORDS = {n: _HINDI_NUMBERS_0_99[n] for n in range(1, 32)}

# English month name -> canonical Hindi month, derived from _HINDI_MONTHS.
_ENGLISH_TO_HINDI_MONTH = {english: hindi for hindi, english in _HINDI_MONTHS.items()}

# Hindi weekday names, indexed by datetime.weekday() (Monday == 0).
_HINDI_WEEKDAYS = (
    "सोमवार",
    "मंगलवार",
    "बुधवार",
    "गुरुवार",
    "शुक्रवार",
    "शनिवार",
    "रविवार",
)


def _spoken_hindi_date(dt: datetime) -> str:
    """Render a date as naturally spoken Hindi words ('1 July' -> 'एक जुलाई')."""
    day = _HINDI_DAY_WORDS.get(dt.day, str(dt.day))
    month = _ENGLISH_TO_HINDI_MONTH.get(dt.strftime("%B"), dt.strftime("%B"))
    return f"{day} {month}"


def _relative_date_reference(today: datetime | None = None) -> str:
    """Build a concrete relative-date table for the prompt.

    Resolving "कल"/"next Monday" to a real date is date arithmetic, which LLMs
    do unreliably. We compute the dates here so the model only has to look them
    up: it speaks the Hindi form shown and passes the DD-MM-YYYY shown to the
    tool, with no arithmetic of its own.
    """
    today = today or datetime.now()

    def line(label: str, dt: datetime) -> str:
        weekday = _HINDI_WEEKDAYS[dt.weekday()]
        return f'- "{label}" = {_spoken_hindi_date(dt)} ({dt.strftime("%d-%m-%Y")}, {weekday})'

    lines = [
        line("आज / today", today),
        line("कल / tomorrow", today + timedelta(days=1)),
        line("परसों / day after tomorrow", today + timedelta(days=2)),
    ]

    # "N दिन बाद" / "in N days" for the next 30 days, so any "X days later" is a
    # direct lookup — this is exactly where the model used to hallucinate dates.
    lines.append("")
    lines.append('Days from today ("N दिन बाद" / "in N days"):')
    for n in range(1, 31):
        dt = today + timedelta(days=n)
        weekday = _HINDI_WEEKDAYS[dt.weekday()]
        lines.append(
            f'- "{_HINDI_NUMBERS_0_99[n]} दिन बाद / in {n} day(s)" = '
            f"{_spoken_hindi_date(dt)} ({dt.strftime('%d-%m-%Y')}, {weekday})"
        )

    # Next occurrence of each weekday (strictly after today, within 1-7 days),
    # so phrases like "अगले सोमवार" / "next Monday" resolve correctly.
    lines.append("")
    lines.append('Next weekday ("अगले <din>" / "next <weekday>"):')
    for offset in range(1, 8):
        dt = today + timedelta(days=offset)
        weekday = _HINDI_WEEKDAYS[dt.weekday()]
        lines.append(
            f'- "अगले {weekday} / next {dt.strftime("%A")}" = '
            f"{_spoken_hindi_date(dt)} ({dt.strftime('%d-%m-%Y')})"
        )
    return "\n".join(lines)


def _hindi_words_to_number(text: str) -> int | None:
    """Parse Hindi number words into an integer ('पाँच हज़ार' -> 5000).

    Returns None if any token is not a recognized number word.
    """
    result = current = 0
    found = False
    for token in str(text).replace(",", " ").split():
        tok = token.strip()
        if not tok or tok in _NUMBER_NOISE:
            continue
        if tok in _HINDI_UNITS:
            current += _HINDI_UNITS[tok]
        elif tok in _HINDI_HUNDRED:
            current = (current or 1) * _HINDI_HUNDRED[tok]
        elif tok in _HINDI_BIG:
            result += (current or 1) * _HINDI_BIG[tok]
            current = 0
        else:
            return None
        found = True
    return result + current if found else None


def normalize_ptp_amount(value) -> str | None:
    """Normalize an amount to a plain integer string ('चार हज़ार रुपए' -> '4000').

    Handles ASCII digits, Devanagari digits, and Hindi number words. Returns
    None when nothing numeric can be extracted.
    """
    if _is_missing(value):
        return None
    text = str(value).strip().replace("₹", "").replace(",", "").translate(_DEV_DIGITS)
    # Already a plain number, e.g. "5000" or "5000.0".
    if re.fullmatch(r"\d+(?:\.\d+)?", text.strip()):
        return str(int(float(text.strip())))
    # Hindi words, e.g. "नौ हज़ार पाँच सौ".
    number = _hindi_words_to_number(text)
    if number is not None:
        return str(number)
    # Last resort: pull any embedded digits ("₹5,000 only" -> "5000").
    digits = "".join(re.findall(r"\d+", text))
    return digits or None


def normalize_ptp_date(value, today: datetime | None = None) -> str | None:
    """Normalize a date to DD-MM-YYYY ('दस जुलाई' -> '10-07-2026').

    Handles existing DD-MM-YYYY input, Devanagari digits, Hindi month names,
    and Hindi day-number words. Missing years default to the current year.
    Returns None when the value cannot be parsed.
    """
    if _is_missing(value):
        return None
    today = today or datetime.now()
    text = str(value).strip().translate(_DEV_DIGITS)

    # Fast path: already a recognizable explicit format.
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
        except ValueError:
            pass

    # Translate Hindi months and day-number words into English/digits.
    for hindi, english in _HINDI_MONTHS.items():
        text = text.replace(hindi, f" {english} ")
    text = " ".join(
        str(_HINDI_UNITS[t]) if t in _HINDI_UNITS else t for t in text.split()
    )

    try:
        # default supplies the year (and any missing parts) for inputs like
        # "10 July"; fuzzy lets us ignore leftover particles such as "को".
        dt = dateutil_parser.parse(text, default=today, dayfirst=True, fuzzy=True)
        return dt.strftime("%d-%m-%Y")
    except (ValueError, OverflowError):
        return None


def migrate_ptp_columns(path: Path = CUSTOMER_DATA_FILE) -> int:
    """Rewrite existing PTP Amount / PTP Date cells into structured English.

    Only the two PTP columns are touched; all other columns and the sheet
    layout are preserved. Returns the number of cells changed.
    """
    try:
        workbook = load_workbook(path)
    except Exception:
        logger.exception("Could not open %s for PTP migration", path)
        return 0

    worksheet = workbook.active

    header_row_idx = None
    columns: dict[str, int] = {}
    for r_idx, sheet_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        labels = {
            str(v).strip().lower(): c
            for c, v in enumerate(sheet_row, start=1)
            if v is not None
        }
        if "customer name" in labels:
            header_row_idx = r_idx
            columns = labels
            break

    amount_col = columns.get("ptp amount")
    date_col = columns.get("ptp date")
    if header_row_idx is None or not (amount_col and date_col):
        logger.error("PTP columns not found in %s; nothing migrated", path)
        return 0

    changed = 0
    for r_idx in range(header_row_idx + 1, worksheet.max_row + 1):
        amount_cell = worksheet.cell(row=r_idx, column=amount_col)
        if not _is_missing(amount_cell.value):
            new_amount = normalize_ptp_amount(amount_cell.value)
            if new_amount is not None and str(amount_cell.value) != new_amount:
                logger.info(
                    "Row %d PTP amount: %r -> %r", r_idx, amount_cell.value, new_amount
                )
                amount_cell.value = new_amount
                changed += 1

        date_cell = worksheet.cell(row=r_idx, column=date_col)
        if not _is_missing(date_cell.value):
            new_date = normalize_ptp_date(date_cell.value)
            if new_date is not None and str(date_cell.value) != new_date:
                logger.info(
                    "Row %d PTP date: %r -> %r", r_idx, date_cell.value, new_date
                )
                date_cell.value = new_date
                changed += 1

    if changed:
        workbook.save(path)
    logger.info("PTP migration complete: %d cell(s) updated in %s", changed, path)
    return changed


def save_promise_to_pay(
    customer_name: str,
    ptp_amount: str,
    ptp_date: str,
    path: Path = CUSTOMER_DATA_FILE,
) -> bool:
    """Write PTP Amount and PTP Date for a customer back into the Excel file.

    Only the two PTP cells of the matching row are touched; every other value
    and the sheet layout are left untouched. Returns True on success.
    """
    # Final safety gate: never persist a partial promise. The tool layer also
    # checks this, but we re-check here so the file can never hold half a PTP.
    if not str(ptp_amount).strip() or not str(ptp_date).strip():
        logger.warning(
            "Refusing to save incomplete PTP (amount=%r, date=%r)", ptp_amount, ptp_date
        )
        return False

    try:
        workbook = load_workbook(path)
        worksheet = workbook.active

        # Locate the header row and the columns we care about, by name, so the
        # code is resilient to column reordering or a title banner above them.
        header_row_idx = None
        columns: dict[str, int] = {}
        for r_idx, sheet_row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            labels = {
                str(v).strip().lower(): c
                for c, v in enumerate(sheet_row, start=1)
                if v is not None
            }
            if "customer name" in labels:
                header_row_idx = r_idx
                columns = labels
                break

        if header_row_idx is None:
            logger.error("No header row with 'Customer Name' found in %s", path)
            return False

        name_col = columns.get("customer name")
        amount_col = columns.get("ptp amount")
        date_col = columns.get("ptp date")
        if not (name_col and amount_col and date_col):
            logger.error("Missing PTP columns in %s (found: %s)", path, list(columns))
            return False

        # Find the row for the loaded customer.
        target_row = None
        for r_idx in range(header_row_idx + 1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=r_idx, column=name_col).value
            if (
                cell_value is not None
                and str(cell_value).strip().lower() == customer_name.strip().lower()
            ):
                target_row = r_idx
                break

        if target_row is None:
            logger.error("Customer %r not found in %s", customer_name, path)
            return False

        worksheet.cell(row=target_row, column=amount_col, value=ptp_amount)
        worksheet.cell(row=target_row, column=date_col, value=ptp_date)
        workbook.save(path)
        logger.info(
            "Saved PTP for %s (row %d): amount=%r date=%r",
            customer_name,
            target_row,
            ptp_amount,
            ptp_date,
        )
        return True
    except Exception:
        logger.exception("Failed to save PTP to %s", path)
        return False


# ---------------------------------------------------------------------------
# Call outcome
#
# Exactly one outcome is recorded per call, written once near the end. If the
# conversation changes direction, the most recent valid outcome wins (the cell
# is simply overwritten). Outcome lives in its own column and its own tool, so
# it never interferes with the PTP Amount / PTP Date cells.
# ---------------------------------------------------------------------------

# The only outcomes the bot may store. Canonical spelling -> lookup key.
VALID_OUTCOMES = [
    "Abusive Customer",
    "Wrong Person",
    "Dont Call Me",
    "Refuse to pay",
    "Already Paid",
    "Call Back",
    "Will Pay",
    "Paying Rightaway",
    "Connect me to Humans",
    "Call Disconnected - No outcome",
    "Discussed - No Outcome",
]
_OUTCOME_LOOKUP = {o.lower(): o for o in VALID_OUTCOMES}


def normalize_outcome(value) -> str | None:
    """Map a model-supplied outcome to its canonical spelling.

    Case-insensitive; returns None for anything not in VALID_OUTCOMES so an
    invented outcome can never be written.
    """
    if _is_missing(value):
        return None
    return _OUTCOME_LOOKUP.get(str(value).strip().lower())


def save_call_outcome(
    customer_name: str,
    outcome: str,
    path: Path = CUSTOMER_DATA_FILE,
) -> bool:
    """Write the final call Outcome for a customer back into the Excel file.

    Only the single Outcome cell of the matching row is touched. Intentionally
    last-write-wins: the final outcome of the call overwrites any earlier one.
    Returns True on success.
    """
    canonical = normalize_outcome(outcome)
    if canonical is None:
        logger.warning("Refusing to save invalid outcome %r", outcome)
        return False

    try:
        workbook = load_workbook(path)
        worksheet = workbook.active

        header_row_idx = None
        columns: dict[str, int] = {}
        for r_idx, sheet_row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            labels = {
                str(v).strip().lower(): c
                for c, v in enumerate(sheet_row, start=1)
                if v is not None
            }
            if "customer name" in labels:
                header_row_idx = r_idx
                columns = labels
                break

        if header_row_idx is None:
            logger.error("No header row with 'Customer Name' found in %s", path)
            return False

        name_col = columns.get("customer name")
        outcome_col = columns.get("outcome")
        if not (name_col and outcome_col):
            logger.error(
                "Missing Outcome column in %s (found: %s)", path, list(columns)
            )
            return False

        target_row = None
        for r_idx in range(header_row_idx + 1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=r_idx, column=name_col).value
            if (
                cell_value is not None
                and str(cell_value).strip().lower() == customer_name.strip().lower()
            ):
                target_row = r_idx
                break

        if target_row is None:
            logger.error("Customer %r not found in %s", customer_name, path)
            return False

        worksheet.cell(row=target_row, column=outcome_col, value=canonical)
        workbook.save(path)
        logger.info(
            "Saved outcome for %s (row %d): %r", customer_name, target_row, canonical
        )
        return True
    except Exception:
        logger.exception("Failed to save outcome to %s", path)
        return False


# ---------------------------------------------------------------------------
# Agent-initiated hangup
#
# Deleting the room ends the call for both web and SIP/telephony participants,
# and fires the session "close" event that runs outcome classification. Kept as
# a module-level function so the end_call tool stays trivially unit-testable
# (tests monkeypatch _hangup instead of touching the LiveKit API).
# ---------------------------------------------------------------------------
# Closing/farewell markers used to detect when Priya has ended the call. These
# are deliberately multi-word, closing-only phrases: a bare "धन्यवाद" also
# appears in the OPENING greeting, so matching on that alone would cut the call
# right after it starts. Romanized variants are included as a fallback.
_FAREWELL_MARKERS = (
    "आपका दिन शुभ",
    "दिन शुभ हो",
    "शुभ दिन",
    "अलविदा",
    "धन्यवाद आपका समय",
    "aapka din shubh",
    "din shubh ho",
    "shubh din",
    "alvida",
)


def _is_farewell(text) -> bool:
    """True if the agent's line is a closing farewell (so the call can end)."""
    if not text:
        return False
    # Strip punctuation (danda "।", commas, etc.) but keep the full Devanagari
    # block U+0900-U+097F -- \w alone drops combining vowel marks (matras),
    # which would mangle the Hindi words. Then collapse whitespace.
    normalized = re.sub(r"[^\w\sऀ-ॿ]", " ", str(text))
    normalized = re.sub(r"\s+", " ", normalized).strip().lower()
    return any(marker in normalized for marker in _FAREWELL_MARKERS)


async def _hangup() -> None:
    """End the call by deleting the room."""
    job_ctx = get_job_context()
    if job_ctx is None:
        logger.warning("Hangup requested but no job context is available")
        return
    try:
        await job_ctx.api.room.delete_room(
            api.DeleteRoomRequest(room=job_ctx.room.name)
        )
        logger.info(">>> HANGUP room=%r deleted", job_ctx.room.name)
    except Exception:
        logger.exception("Failed to delete room during hangup")


# ---------------------------------------------------------------------------
# Post-call outcome classification
#
# The outcome is NOT recorded by the voice agent via a tool. Asking the speaking
# model to emit a tool call inside its closing turn caused the tool-call syntax
# to leak into the customer transcript (and the call was never dispatched).
#
# Instead, once the call ends (including on disconnect), we run a single,
# non-voice classification over the transcript and persist exactly one outcome.
# This keeps outcome selection completely off the voice path, so nothing
# internal can ever reach the customer.
# ---------------------------------------------------------------------------

_OUTCOME_DEFINITIONS = """\
- "Abusive Customer" — the customer is abusive, threatening, or uses offensive language.
- "Wrong Person" — the person reached is not the customer (wrong number, or no such person / says it is not them).
- "Dont Call Me" — the customer asks not to be called again.
- "Refuse to pay" — the customer clearly refuses to pay and makes no commitment.
- "Already Paid" — the customer says they have already paid, or that the account is settled.
- "Call Back" — the customer (or whoever answered) asks to be called back later, or says the customer is busy / not available right now.
- "Will Pay" — the customer commits to pay on a future date.
- "Paying Rightaway" — the customer says they are paying now / immediately.
- "Connect me to Humans" — the customer asks to be transferred to a human agent.
- "Call Disconnected - No outcome" — the call dropped before any resolution was reached.
- "Discussed - No Outcome" — the matter was discussed but none of the above applies.
"""


def _match_outcome_in_text(text) -> str | None:
    """Map the classifier's free-form reply to one canonical outcome label."""
    if _is_missing(text):
        return None
    cleaned = str(text).strip().strip("\"'`.").strip()
    exact = normalize_outcome(cleaned)
    if exact is not None:
        return exact
    # The model may wrap the label in a sentence; find the first label present.
    lowered = str(text).lower()
    for outcome in VALID_OUTCOMES:
        if outcome.lower() in lowered:
            return outcome
    return None


def _transcript_from_history(history: ChatContext) -> str:
    """Render the call's user/assistant turns into a plain transcript."""
    lines: list[str] = []
    for item in history.items:
        role = getattr(item, "role", None)
        if role not in ("user", "assistant"):
            continue
        text = (item.text_content or "").strip()
        if not text:
            continue
        lines.append(f"{'Customer' if role == 'user' else 'Agent'}: {text}")
    return "\n".join(lines)


def _is_valid_hindi_text(text: str) -> bool:
    """Check if transcribed text looks like Hindi/Hinglish, not garbage.

    Garbage indicators: text in Spanish, French, random Latin scripts that
    aren't English + Hindi. Returns False if it looks like a bad STT output.
    """
    if not text or len(text.strip()) < 2:
        return True  # too short to judge, allow it

    # Count scripts
    devanagari_chars = sum(1 for c in text if "\u0900" <= c <= "\u097f")
    latin_chars = sum(1 for c in text if c.isascii() and c.isalpha())
    total_alpha = devanagari_chars + latin_chars

    if total_alpha == 0:
        return True  # no letters, probably just punctuation/numbers

    # Hindi/Hinglish should have at least 20% Devanagari OR be mostly English
    # If it's a mix with very little Devanagari, it's likely garbage (Spanish etc)
    devanagari_ratio = devanagari_chars / total_alpha
    latin_ratio = latin_chars / total_alpha

    if total_alpha > 5 and devanagari_ratio < 0.15 and latin_ratio < 0.85:
        # Lots of text, almost no Devanagari, and not pure English either
        logger.warning(
            "Garbage STT detected: %r (%.1f%% Devanagari, %.1f%% Latin)",
            text,
            devanagari_ratio * 100,
            latin_ratio * 100,
        )
        return False

    return True


# ---------------------------------------------------------------------------
# STT noise filtering (filler sounds + duplicate finals)
#
# These are applied in Assistant.stt_node, which is the only reliable place to
# drop an utterance BEFORE it becomes a user turn: there is no "user_speech"
# session event, so an .on(...) handler can never block one. Dropping the
# FINAL_TRANSCRIPT here means the utterance never reaches the transcript, the
# LLM, or preemptive generation.
# ---------------------------------------------------------------------------

# Filler sounds with any elongation: "hmm", "uhh", "uhmm", "ahh", "err", ...
_FILLER_TOKEN_RE = re.compile(
    r"u+h+m*|u+m+|u+h*|h+m+|h+|m+|a+h+|e+r+|e+h+|o+h+|huh|hmm+"
)
# Devanagari spellings the STT may emit for the same sounds. Deliberately
# excludes meaningful words like "हाँ" (yes) / "ना" (no).
_DEVANAGARI_FILLERS = {
    "हम्म",
    "हम्म्म",
    "हँ",
    "अं",
    "अँ",
    "उम",
    "उह",
    "अह",
    "आह",
    "ऊँ",
    "एं",
}
# Window within which an identical final transcript is treated as an STT
# duplicate rather than the customer genuinely repeating themselves.
_STT_DEDUP_WINDOW_S = 6.0


def _is_filler_only(text: str) -> bool:
    """True if the utterance is nothing but filler sounds ('hmm', 'उह')."""
    if _is_missing(text):
        return False
    tokens = re.findall(r"[a-zऀ-ॿ]+", str(text).lower())
    if not tokens:
        return False
    for tok in tokens:
        if tok in _DEVANAGARI_FILLERS or _FILLER_TOKEN_RE.fullmatch(tok):
            continue
        return False
    return True


def _dedup_key(text: str) -> str:
    """Normalize an utterance (lowercase, punctuation-stripped) for dedup compare."""
    cleaned = re.sub(r"[^\w]+", " ", str(text).lower())
    return re.sub(r"\s+", " ", cleaned).strip()


async def classify_and_save_outcome(history: ChatContext, customer_name: str) -> None:
    """Classify the finished call into exactly one outcome and persist it."""
    try:
        transcript = _transcript_from_history(history)
        if not transcript:
            # Nothing was exchanged — the call dropped before any resolution.
            save_call_outcome(customer_name, "Call Disconnected - No outcome")
            return

        ctx = ChatContext.empty()
        ctx.add_message(
            role="system",
            content=(
                "You classify the outcome of a debt-collection phone call. "
                "Read the transcript and pick the single label that best describes "
                "how the call ended. If the customer changed direction during the "
                "call, use the most recent valid outcome. Reply with the label "
                "text ONLY, nothing else.\n\n" + _OUTCOME_DEFINITIONS
            ),
        )
        ctx.add_message(role="user", content=f"Transcript:\n{transcript}\n\nOutcome:")

        # A small, fast model — this is a trivial one-label classification, so we
        # avoid the heavy voice model. temperature=0 makes it deterministic and
        # max_completion_tokens caps it to just the label, cutting latency to ~1s.
        classifier = inference.LLM(
            model="openai/gpt-4o-mini",
            extra_kwargs={"temperature": 0, "max_completion_tokens": 16},
        )
        raw = ""
        try:
            async with classifier.chat(chat_ctx=ctx) as stream:
                async for chunk in stream:
                    if chunk.delta and chunk.delta.content:
                        raw += chunk.delta.content
        finally:
            await classifier.aclose()

        outcome = _match_outcome_in_text(raw) or "Discussed - No Outcome"
        logger.info(
            ">>> OUTCOME-CLASSIFY customer=%r raw=%r -> %r",
            customer_name,
            raw.strip(),
            outcome,
        )
        save_call_outcome(customer_name, outcome)
    except Exception:
        logger.exception("Outcome classification failed for %r", customer_name)


# NOTE: The strict identity-verification flow was disabled in favor of a simple
# name confirmation (see the "IDENTITY CHECK (SIMPLE)" block in the prompt).
# The previous behavior is kept here, commented out, so it can be restored:
#
#   IDENTITY VERIFICATION
#   Before identity verification:
#     * Do not reveal loan amount / due amount / due date / account status.
#     * Do not reveal any customer-specific information.
#   If identity cannot be verified, politely end the call.
class Assistant(Agent):
    def __init__(self) -> None:
        # Resolve relative dates ("कल", "next Monday") in code and hand the model
        # a concrete lookup table, so it never has to do date arithmetic itself.
        now = datetime.now()
        relative_dates = _relative_date_reference(now)
        super().__init__(
            llm=inference.LLM(model="openai/gpt-4.1-mini"),
            instructions=textwrap.dedent(f"""
You are Priya, an experienced debt resolution specialist calling on behalf of a financial services company.

================================================
CUSTOMER INFORMATION
====================

{CUSTOMER_CONTEXT}

Use this information naturally once the person confirms they are {LOADED_CUSTOMER_NAME}.

Refer to the lender by the Lender Name on file (for example, "आपके HDFC Bank लोन को लेकर"). Only mention the lender after the person confirms they are the customer, and never assume a lender that is not listed above.

Do not reveal loan amount, due amount, due date, or any account details until the person confirms they are {LOADED_CUSTOMER_NAME}.

================================================
IDENTITY CHECK (SIMPLE)
=======================

Open by confirming, by name, that you are speaking to the customer, for example:
"नमस्ते, क्या मेरी बात {LOADED_CUSTOMER_NAME} जी से हो रही है?"

Then:

* If the person confirms they are {LOADED_CUSTOMER_NAME} (for example "हाँ", "जी हाँ", "speaking", "yes") → continue the call normally.
* If the person is NOT {LOADED_CUSTOMER_NAME} — wrong number, no such person, or they say it is not them → treat as WRONG PERSON. Do not share any account details. Apologize briefly and end the call.
* If the person knows {LOADED_CUSTOMER_NAME} but says they are busy, not available, cannot talk right now, or asks you to call later → treat as CALLBACK. Politely say you will call back later and end the call. Do not share account details.

A simple name confirmation is enough. Do not ask any additional verification questions (no date of birth, no account number, no OTP).

================================================
LANGUAGE
================================================



* Default language: Hinglish.

* Default script:
  - Use Devanagari for Hindi words.
  - Use English for common BFSI and business terminology.
  - Do not ask language preference. Switch permanently to English only when explicitly requested.

* Hindi customer:
  - Respond primarily in Hindi written in Devanagari.
  - Use natural english BFSI terminology where appropriate.

* English customer:

  - Respond entirely in English.

* Hinglish customer:
  - Use natural Indian Hinglish.
  - Write Hindi words in Devanagari and business terms in English.

Examples:

Correct:
"नमस्ते राहुल जी।"

Correct:
"आपके personal loan account में payment due है।"

Correct:
"क्या आप payment आज कर पाएंगे?"

Correct:
"मैं callback request बना देती हूँ।"

Correct:
"क्या आप human agent से बात करना चाहेंगे?"

Avoid overly formal Hindi.

Incorrect:
"मैं पुनर्भुगतान संग्रहण विभाग से बोल रही हूँ।"

Incorrect:
"आपका ऋण खाता बकाया है।"

================================================
BFSI TERMINOLOGY (IMPORTANT)
================================================

Always use the standard English BFSI term, written in Latin script, for the
words below. NEVER use their pure/literary Hindi translations. This applies
everywhere, including your standard and confirmation lines.

* "payment" — never भुगतान
* "amount" — never राशि (do not use राशि at all; just say "payment")
* "outstanding" / "due" / "overdue" — never बकाया
* "loan" — never ऋण
* "EMI" / "installment" — never किस्त / क़िस्त
* "due date" — never देय तिथि / नियत तिथि
* "confirm" — never पुष्टि
* "verify" — never सत्यापन
* "review" — never समीक्षा
* "settlement" — never समझौता
* "account" — never लेखा

When asking how much the customer will pay, ask "कितनी payment कर पाएंगे?" —
never insert राशि (do not say "कितनी राशि payment कर पाएंगे").

Correct:
"आप कितनी payment कर पाएंगे?"

Incorrect:
"आप कितनी राशि payment कर पाएंगे?"

Correct:
"आपका payment अभी outstanding है।"

Incorrect:
"आपकी राशि अभी बकाया है।"

* Do not ask the customer which language they prefer.
* Start every call in natural Hindi/Hinglish.
* If language is unclear, continue in Hindi/Hinglish.
* Never write Hindi in Roman script.
* Speak only Hindi, Hinglish, or English.

Never speak Spanish, French, German, or any other language.

================================================
HINDI STYLE (FOR SARVAM TTS)
============================

All Hindi must be optimized for natural Sarvam text-to-speech pronunciation.

* Use short, conversational spoken Hindi.
* Avoid overly formal or literary Hindi.
* Prefer everyday spoken language.
* Use correct Devanagari spelling and matras.
* Keep sentences short so speech sounds natural on a phone call.

================================================
SPEECH FORMAT
=============

When speaking to customers:

* Use natural conversational language.
* Keep replies short.
* Ask only one question at a time.
* Never use digits when speaking money amounts or dates.
* Speak amounts and dates in naturally spoken Hindi words when using Hindi.
* Read OTPs and reference numbers digit by digit.
* Pause and wait for the customer's response.

Examples:

₹18,750 → अठारह हज़ार सात सौ पचास रुपये

₹2,00,000 → दो लाख रुपये

15 June → पंद्रह जून

================================================
TOOL FORMAT
===========

When calling tools:

* Amount must be numeric only.
* Date must be DD-MM-YYYY.
* Never pass Hindi text to tools.
* Spoken language and tool arguments are separate concerns.

Example:

Customer:
"मैं पाँच हज़ार रुपये पच्चीस जून को दे दूँगा"

Tool:
amount="5000"
date="25-06-2026"

Today's date is {now.strftime("%d-%m-%Y")} ({_HINDI_WEEKDAYS[now.weekday()]}).

RELATIVE DATES — use this exact table. You are FORBIDDEN from calculating any
date yourself; doing so has produced wrong dates. Every relative day the customer
can say is already in the table below. Find the matching row, SPEAK the Hindi
form shown, and pass the DD-MM-YYYY shown to the tool.

This explicitly includes "N दिन बाद" / "N days later" (for example "दस दिन बाद",
"तीन दिन बाद"): look up the exact "N दिन बाद" row — NEVER add the days yourself.

{relative_dates}

When confirming a relative date with the customer, always state the date from the
matching row (e.g. "दस दिन बाद से मतलब <table date>") — never a date you computed.
For "अगले हफ्ते" / "next week", confirm a specific day from the table first. For a
date not in the table (for example "महीने की 5 तारीख" / "5th of the month"), use
the explicit day with the current or next month as the customer means, in
DD-MM-YYYY.

================================================
INDIAN DATE INTERPRETATION
================================================

Customers may express payment dates informally.

Examples:

"कल"
"परसों"
"अगले सोमवार"
"अगले हफ्ते"
"महीने की 5 तारीख"
"salary आते ही"

Interpret these naturally in conversation.

Before recording a Promise To Pay, always confirm the exact payment date if there is any ambiguity.

Example (uses the table above — "कल" resolves to {_spoken_hindi_date(now + timedelta(days=1))}):

Customer:
"कल कर दूँगा।"

Agent:
"जी, confirm कर दूँ — क्या आप {_spoken_hindi_date(now + timedelta(days=1))} को payment करेंगे?"

Do not use pure Hindi words like पुष्टि, use words like confirm, verify instead.

Only record a Promise To Pay after the customer confirms the specific date.

================================================
TOOL FAILURE HANDLING
=====================

If a tool fails:

* Do not tell the customer that a tool failed.
* Continue the conversation normally.
* Acknowledge the customer's commitment.
* Close the call professionally.

================================================
VOICE STYLE
===========

You are speaking on a live phone call.

Sound like an experienced collections agent, not a chatbot.

* Calm
* Professional
* Human
* Conversational

Keep replies short.

Usually one or two sentences.

Do not sound scripted.

Do not sound like customer support documentation.

You may occasionally use natural fillers such as:

* जी
* समझ गई
* ठीक है
* बिलकुल

Do not use pure Hindi words like pushti or samjhauta; use natural Hinglish instead.

Do not overuse them.

================================================
CONVERSATION MEMORY
===================

Maintain conversation state throughout the call.

Track internally whether the customer has already:

* verified identity
* confirmed they are the correct person
* provided a payment amount
* provided a payment date
* explained inability to pay
* provided a dispute reason
* requested a callback
* requested settlement
* requested escalation

Never ask again for information that has already been collected unless clarification is genuinely required.

Treat previously collected information as known for the remainder of the call.

Always move the conversation forward.

================================================
ANTI-REPETITION
===============

This is extremely important.

Never repeat:

* introductions
* verification requests
* account explanations
* payment explanations
* escalation messages
* callback requests
* previously answered questions

If a customer has already answered a question:

* acknowledge the answer
* build on it
* move forward

Do not ask the same question using different wording.

Bad Example:

Agent: "When can you make the payment?"

Customer: "Next Monday."

Agent: "What date can you make the payment?"

Wrong.

Good Example:

Agent: "Understood. How much will you be able to pay on Monday?"

Correct.

If the customer repeats the same answer multiple times:

* Do not repeat the same question.
* Gather the next required information.
* Offer an alternative.
* Move toward resolution.
* Escalate if appropriate.

If a topic has already been discussed and no new information is available, do not revisit it.

Every response should do at least one of the following:

* collect new information
* confirm a commitment
* resolve an objection
* progress toward closure
* progress toward escalation

Avoid conversational loops at all costs.

If no new information is being obtained after two exchanges on the same topic, either move to the next step of the conversation or escalate when appropriate.

================================================
INTERRUPTION HANDLING
=====================

If the customer interrupts:

* Address the interruption immediately.
* Do not continue the previous sentence.
* Do not force the conversation back to the interrupted script.
* Respond naturally and then continue.

================================================
CALL FLOW
=========

1. Introduce yourself as Priya.
2. Confirm the correct person.
3. Verify identity.
4. Discuss the account.
5. Work toward resolution.
6. Close or escalate.

This flow is guidance, not a script.

Sound natural.

================================================
PAYMENT RESOLUTION PRINCIPLES
=============================

When discussing repayment:

* Understand the customer's intent first.
* Work toward a specific commitment.
* Prefer specific dates over vague promises.
* Prefer specific amounts over general willingness to pay.

Tentative statements are NOT commitments.

Examples:

* Maybe
* I'll try
* Let's see
* Hopefully
* Should be able to

Do not treat these as confirmed promises.

================================================
OUTCOMES
========

PROMISE TO PAY

Collect:

* payment amount
* payment date

Confirm both once.

Only after the customer clearly confirms both:

* call record_promise_to_pay
* pass structured tool values
* close politely

Keep speaking naturally in Hindi/Hinglish, but pass structured English values to the tool:

* amount: digits only in rupees
* date: DD-MM-YYYY

Do not invent values.

Use only information explicitly provided by the customer.

UNABLE TO PAY

Explore one realistic future payment date.

REFUSAL TO PAY

Ask the reason once.

Attempt one resolution.

If unresolved, escalate.

ALREADY PAID

Collect payment date and reference number.

Mark for review.

DISPUTE

Collect dispute reason.

Escalate for review.

CALLBACK

Collect callback date and time.

Confirm once.

SETTLEMENT REQUEST

Collect request details.

Escalate.

WRONG PARTY

Reveal no account information.

Apologize briefly and close with the farewell so the call ends.

================================================
CALL CLOSING
============

Once a clear outcome has been reached:

* Do not reopen negotiation.
* Do not ask new questions.
* Summarize the next action briefly.
* End politely.

Keep closing statements under two sentences.

When the conversation is finished (an outcome has been reached, or it is a wrong
party, or there is nothing more to do), end with a short farewell. Your final
sentence MUST be a clear goodbye, ending with "धन्यवाद, आपका दिन शुभ हो।". The
call ends automatically right after you say it, so say it only when you are truly
done — do not say it mid-conversation.

NEVER speak, type, read aloud, or output any function name, code, JSON, or tool
syntax (for example never produce "end_call", "functions.end_call", or
"{{ outcome_reached: ... }}"). Speak only natural Hindi to the customer. The call
ending is handled automatically from your spoken farewell — you do not call any
tool to end it.

================================================
ESCALATION
==========

Escalate when:

* repeated refusal
* supervisor requested
* settlement requested
* legal threats mentioned
* dispute requires review

When escalating, say:

"मैं आपका केस सीनियर टीम को रिव्यू के लिए फॉरवर्ड कर रही हूँ।"

Do not continue negotiation afterward.

================================================
IMPORTANT
=========

Never mention being AI.

Never use markdown in speech.

Never use emojis.

Never congratulate the customer or use celebratory words such as "बधाई",
"मुबारक", or "congratulations". A payment is pending — there is nothing to
celebrate. Open with a simple "नमस्ते" and get to the point directly, warmly,
and politely.

Never speak the ₹ symbol or any digits for money. Always state the Outstanding
Amount exactly as written in CUSTOMER INFORMATION, in Hindi words (for example
"नौ हज़ार पाँच सौ रुपये", never "₹9500" or "9500").

Stay calm.

Stay human.

Stay conversational.

Your goal is to achieve a clear resolution and collect the next best action.

"""),
        )
        # Last committed user utterance, for the turn-level duplicate backstop
        # in on_user_turn_completed (complements the stt_node final-dedup).
        self._last_user_key = ""
        self._last_user_at = 0.0

    async def on_user_turn_completed(self, turn_ctx, new_message) -> None:
        """Drop a user turn that exactly repeats the previous one.

        A second safety net behind stt_node: if the STT still surfaces the same
        utterance twice in quick succession, we discard the duplicate turn so it
        never reaches the transcript or triggers a second response.
        """
        text = (new_message.text_content or "").strip()
        key = _dedup_key(text)
        now = time.monotonic()
        if (
            key
            and key == self._last_user_key
            and (now - self._last_user_at) < _STT_DEDUP_WINDOW_S
        ):
            logger.info(">>> DUPLICATE TURN dropped: %r", text)
            raise StopResponse()
        self._last_user_key, self._last_user_at = key, now

    async def stt_node(self, audio, model_settings):
        """Filter the STT stream before it becomes a user turn.

        Drops three kinds of final transcripts so they never reach the
        transcript, the LLM, or preemptive generation:
          * garbage (non-Hindi/English STT misfires),
          * filler-only utterances ("hmm", "उह"),
          * back-to-back duplicate finals (the same line recognized twice).
        Everything else (including interim transcripts) passes through
        untouched, so VAD and turn detection are unaffected.
        """
        base = Agent.default.stt_node(self, audio, model_settings)
        if inspect.isawaitable(base):
            base = await base
        if base is None:
            return

        last = {"key": "", "at": 0.0}
        async for event in base:
            if (
                event.type == stt.SpeechEventType.FINAL_TRANSCRIPT
                and event.alternatives
            ):
                text = event.alternatives[0].text or ""

                if not _is_valid_hindi_text(text):
                    logger.error(">>> GARBAGE STT dropped: %r", text)
                    continue
                if _is_filler_only(text):
                    logger.info(">>> FILLER STT dropped: %r", text)
                    continue

                key = _dedup_key(text)
                now = time.monotonic()
                if (
                    key
                    and key == last["key"]
                    and (now - last["at"]) < _STT_DEDUP_WINDOW_S
                ):
                    logger.info(">>> DUPLICATE STT dropped: %r", text)
                    continue
                last["key"], last["at"] = key, now

            yield event

    @function_tool
    async def record_promise_to_pay(
        self,
        context: RunContext,
        amount: str,
        date: str,
    ) -> str:
        """
        Save a confirmed Promise To Pay.

        Call this only after the customer confirms both a payment amount
        and payment date.

        Pass STRUCTURED ENGLISH values, never Hindi text:
          - amount: digits only, in rupees, e.g. "5000" (not "पाँच हज़ार रुपये").
          - date: DD-MM-YYYY, e.g. "25-06-2026". Resolve relative dates such as
            "next Monday" to an absolute date with the year before calling.

        Do not invent values.
        """
        logger.info(
            ">>> PTP-TRACE step=1 tool_dispatched amount=%r date=%r loaded_customer=%r",
            amount,
            date,
            LOADED_CUSTOMER_NAME,
        )

        if not amount.strip() or not date.strip():
            logger.warning(
                "record_promise_to_pay called with missing amount/date; ignoring"
            )
            return "Promise to pay not saved."

        # Normalize defensively so Hindi text can never reach Excel, even if the
        # model passes spoken-form values.
        clean_amount = normalize_ptp_amount(amount)
        clean_date = normalize_ptp_date(date)
        if clean_amount is None or clean_date is None:
            logger.warning(
                "Could not normalize PTP (amount=%r -> %r, date=%r -> %r)",
                amount,
                clean_amount,
                date,
                clean_date,
            )
            return "Promise to pay not saved: amount/date not understood."

        # Speak a short acknowledgment the instant the tool runs, so the customer
        # hears a response instead of silence while the model's spoken
        # confirmation (a second LLM round-trip) is still being generated. Placed
        # after validation so we never acknowledge an unusable promise.
        context.session.say("जी, ठीक है।")

        # Run the blocking Excel write off the event loop so audio is never
        # stalled by file I/O.
        saved = await asyncio.to_thread(
            save_promise_to_pay,
            LOADED_CUSTOMER_NAME,
            clean_amount,
            clean_date,
        )

        if saved:
            return f"PTP saved: {clean_amount} on {clean_date}."

        return "PTP save failed."


server = AgentServer()


def prewarm(proc: JobProcess):
    proc.userdata["vad"] = silero.VAD.load(
        min_speech_duration=0.045,  # 100
        prefix_padding_duration=0.6,  # 200 ms speech pad
        min_silence_duration=0.65,
        activation_threshold=0.26,  # 200 ms silence pad
    )


server.setup_fnc = prewarm


@server.rtc_session(agent_name="my-agent")
async def my_agent(ctx: JobContext):

    ctx.log_context_fields = {"room": ctx.room.name}

    session = AgentSession(
        stt=sarvam.STT(
            model="saaras:v3",
            language="hi-IN",
            mode="transcribe",
            high_vad_sensitivity=True,
        ),
        tts=sarvam.TTS(
            target_language_code="hi-IN",
            model="bulbul:v3",
            speaker="ishita",
            pace=1.1,
        ),
        turn_detection=TurnDetector(version="v1-mini"),
        vad=ctx.proc.userdata["vad"],
        preemptive_generation=True,
    )

    # NOTE: garbage/filler/duplicate STT filtering now lives in
    # Assistant.stt_node (see that method). The previous @session.on("user_speech")
    # handler was a no-op — LiveKit has no "user_speech" event — so it never
    # actually blocked anything; stt_node is the layer that does.

    # End the call once Priya delivers her closing farewell. We detect the
    # farewell in her transcript (deterministically) rather than via a tool call:
    # asking the speaking model to emit a tool call on its closing turn makes the
    # raw call syntax leak into the spoken transcript (and TTS reads it aloud).
    # We wait for the farewell to finish playing before deleting the room.
    farewell_state = {"pending": False, "done": False}

    async def _do_hangup():
        if farewell_state["done"]:
            return
        farewell_state["done"] = True
        # Deleting the room ends both web and SIP sessions and triggers the
        # "close" handler that classifies and saves the outcome.
        await _hangup()

    async def _end_after_farewell():
        # Wait for the farewell utterance to finish playing, then drop the call.
        speech = session.current_speech
        if speech is not None:
            try:
                await speech.wait_for_playout()
            except Exception:
                logger.exception("wait_for_playout failed during farewell hangup")
            await _do_hangup()
        # If the speech handle isn't up yet, the agent_state_changed backstop
        # below fires the hangup once Priya finishes speaking (returns to idle).

    @session.on("conversation_item_added")
    def _detect_farewell(ev):
        item = ev.item
        if getattr(item, "role", None) != "assistant" or farewell_state["pending"]:
            return
        if _is_farewell(item.text_content or ""):
            logger.info(">>> FAREWELL detected; ending call after playout")
            farewell_state["pending"] = True
            asyncio.create_task(_end_after_farewell())

    @session.on("agent_state_changed")
    def _hangup_when_done_speaking(ev):
        # Backstop: once the farewell has been flagged and Priya stops speaking,
        # end the call. Guarded by _do_hangup so it never double-fires.
        if farewell_state["pending"] and ev.new_state in ("listening", "idle"):
            asyncio.create_task(_do_hangup())

    # Record the call outcome off the voice path, so no tool-call text can leak
    # to the customer. We kick it off on the session "close" event (fires ~2s
    # after hangup) instead of waiting for job shutdown, which LiveKit delays
    # ~20s for room linger. The shutdown callback is only a safety net that
    # awaits the already-running task (no double write, guaranteed completion).
    outcome_state: dict[str, asyncio.Task | None] = {"task": None}

    @session.on("close")
    def _on_session_close(_ev):
        if outcome_state["task"] is None:
            outcome_state["task"] = asyncio.create_task(
                classify_and_save_outcome(session.history, LOADED_CUSTOMER_NAME)
            )

    async def _ensure_outcome_saved():
        task = outcome_state["task"]
        if task is None:
            # close never fired (unusual teardown) — classify now.
            await classify_and_save_outcome(session.history, LOADED_CUSTOMER_NAME)
        elif not task.done():
            await task

    ctx.add_shutdown_callback(_ensure_outcome_saved)

    await session.start(
        agent=Assistant(),
        room=ctx.room,
        room_options=room_io.RoomOptions(
            audio_input=room_io.AudioInputOptions(
                noise_cancellation=ai_coustics.audio_enhancement(
                    model=ai_coustics.EnhancerModel.QUAIL_VF_S
                ),
            ),
        ),
    )

    await ctx.connect()

    # FIRST SPEECH (STRICT CONTROL)
    await session.generate_reply(
        instructions=textwrap.dedent(f"""
Start the call in Hindi.

Introduce yourself as Priya from the Alpha financial services company.

Ask, by name, if you are speaking to {LOADED_CUSTOMER_NAME}
(for example: "नमस्ते, क्या मेरी बात {LOADED_CUSTOMER_NAME} जी से हो रही है?").

Keep it short, natural, and conversational.

Open with a simple "नमस्ते". Do NOT congratulate or use words like "बधाई" /
"मुबारक" / "congratulations" — there is nothing to celebrate.

Do NOT repeat introduction again after this.
Do NOT ask multiple questions.
""")
    )


if __name__ == "__main__":
    cli.run_app(server)
