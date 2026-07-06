import asyncio
import inspect
import logging
import math
import os
import re
import textwrap
import time
from datetime import datetime, timedelta
from pathlib import Path

from dateutil import parser as dateutil_parser
from dotenv import load_dotenv
from livekit import api
from livekit.agents import (
    Agent,
    AgentServer,
    AgentSession,
    JobContext,
    JobProcess,
    MetricsCollectedEvent,
    RunContext,
    StopResponse,
    cli,
    function_tool,
    get_job_context,
    inference,
    metrics,
    room_io,
    stt,
)
from livekit.agents.inference import TurnDetector
from livekit.agents.llm import ChatContext
from livekit.plugins import ai_coustics, aws, openai, sarvam, silero
from openpyxl import load_workbook

logger = logging.getLogger("agent")

load_dotenv(".env.local")

# Dispatch name this worker registers under. Override via the AGENT_NAME env var
# so a local `dev` worker (e.g. AGENT_NAME=my-agent-dev) never collides with the
# deployed cloud agent (default "my-agent") — dispatches route only to the
# matching name. telephony/dial.py reads the same var so it targets the right one.
AGENT_NAME = os.environ.get("AGENT_NAME", "my-agent")

# Excel source for customer details, resolved relative to the project root so it
# works regardless of the current working directory.
CUSTOMER_DATA_FILE = Path(__file__).parent.parent / "data" / "Customers.xlsx"

# Name of the customer loaded for this call. Used to locate the row to update
# when persisting a Promise To Pay. We assume one customer is loaded per call.
LOADED_CUSTOMER_NAME = ""


def _is_missing(value) -> bool:
    """Return True for None, NaN, or blank cells."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


# Hindi words for 0-99. Indian numbers in this range are irregular (each has a
# unique word), so they are listed explicitly. Used to SPEAK amounts and dates
# fully in Hindi (Devanagari) — never half-English like "18 hazaar 750".
_HINDI_NUMBERS_0_99 = {
    0: "शून्य",
    1: "एक",
    2: "दो",
    3: "तीन",
    4: "चार",
    5: "पाँच",
    6: "छह",
    7: "सात",
    8: "आठ",
    9: "नौ",
    10: "दस",
    11: "ग्यारह",
    12: "बारह",
    13: "तेरह",
    14: "चौदह",
    15: "पंद्रह",
    16: "सोलह",
    17: "सत्रह",
    18: "अठारह",
    19: "उन्नीस",
    20: "बीस",
    21: "इक्कीस",
    22: "बाईस",
    23: "तेईस",
    24: "चौबीस",
    25: "पच्चीस",
    26: "छब्बीस",
    27: "सत्ताईस",
    28: "अट्ठाईस",
    29: "उनतीस",
    30: "तीस",
    31: "इकतीस",
    32: "बत्तीस",
    33: "तैंतीस",
    34: "चौंतीस",
    35: "पैंतीस",
    36: "छत्तीस",
    37: "सैंतीस",
    38: "अड़तीस",
    39: "उनतालीस",
    40: "चालीस",
    41: "इकतालीस",
    42: "बयालीस",
    43: "तैंतालीस",
    44: "चौवालीस",
    45: "पैंतालीस",
    46: "छियालीस",
    47: "सैंतालीस",
    48: "अड़तालीस",
    49: "उनचास",
    50: "पचास",
    51: "इक्यावन",
    52: "बावन",
    53: "तिरपन",
    54: "चौवन",
    55: "पचपन",
    56: "छप्पन",
    57: "सत्तावन",
    58: "अट्ठावन",
    59: "उनसठ",
    60: "साठ",
    61: "इकसठ",
    62: "बासठ",
    63: "तिरसठ",
    64: "चौंसठ",
    65: "पैंसठ",
    66: "छियासठ",
    67: "सड़सठ",
    68: "अड़सठ",
    69: "उनहत्तर",
    70: "सत्तर",
    71: "इकहत्तर",
    72: "बहत्तर",
    73: "तिहत्तर",
    74: "चौहत्तर",
    75: "पचहत्तर",
    76: "छिहत्तर",
    77: "सतहत्तर",
    78: "अठहत्तर",
    79: "उन्यासी",
    80: "अस्सी",
    81: "इक्यासी",
    82: "बयासी",
    83: "तिरासी",
    84: "चौरासी",
    85: "पचासी",
    86: "छियासी",
    87: "सत्तासी",
    88: "अट्ठासी",
    89: "नवासी",
    90: "नब्बे",
    91: "इक्यानवे",
    92: "बानवे",
    93: "तिरानवे",
    94: "चौरानवे",
    95: "पंचानवे",
    96: "छियानवे",
    97: "सत्तानवे",
    98: "अट्ठानवे",
    99: "निन्यानवे",
}


def _number_to_hindi_words(amount: int) -> str:
    """Spell an integer in Indian-system Hindi words (18750 -> 'अठारह हज़ार सात सौ पचास')."""
    if amount < 0:
        return "माइनस " + _number_to_hindi_words(-amount)
    if amount == 0:
        return _HINDI_NUMBERS_0_99[0]

    crore, rem = divmod(amount, 10_000_000)
    lakh, rem = divmod(rem, 100_000)
    thousand, rem = divmod(rem, 1_000)
    hundred, rem = divmod(rem, 100)  # rem is now 0-99

    parts: list[str] = []
    if crore:
        # crore can exceed 99 for very large values; recurse so it still reads.
        parts += [_number_to_hindi_words(crore), "करोड़"]
    if lakh:
        parts += [_HINDI_NUMBERS_0_99[lakh], "लाख"]
    if thousand:
        parts += [_HINDI_NUMBERS_0_99[thousand], "हज़ार"]
    if hundred:
        parts += [_HINDI_NUMBERS_0_99[hundred], "सौ"]
    if rem:
        parts.append(_HINDI_NUMBERS_0_99[rem])
    return " ".join(parts)


def _to_rupees(value) -> int | None:
    """Parse a rupee amount into an int, tolerating the formatted strings the
    Excel actually stores (e.g. '₹9,500', 'Rs. 9,500', '9500.0'). Returns None if
    there are no digits to parse."""
    if _is_missing(value):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    # Match the first number (e.g. "1,00,000" or "9500.0"), ignoring surrounding
    # "₹", "Rs.", spaces — so a stray "." in "Rs." can't corrupt the parse.
    match = re.search(r"\d[\d,]*(?:\.\d+)?", str(value))
    if not match:
        return None
    try:
        return int(float(match.group().replace(",", "")))
    except (TypeError, ValueError):
        return None


def _format_indian_amount(value) -> str:
    """Format a rupee amount fully in spoken Hindi (18750 -> 'अठारह हज़ार सात सौ पचास रुपये')."""
    if _is_missing(value):
        return "N/A"
    amount = _to_rupees(value)
    if amount is None:
        return str(value).strip()
    return f"{_number_to_hindi_words(amount)} रुपये"


def _recommended_min_payment(value) -> str:
    """Half the outstanding amount, rounded to the NEAREST thousand, spoken in
    Hindi (e.g. 9500 -> half 4750 -> "पाँच हज़ार रुपये").

    This is the amount the agent recommends when the customer asks how much to
    pay. Precomputed here rather than by the LLM, matching how relative dates are
    handled — the model must never do the arithmetic itself. 'N/A' if the amount
    is missing or unparseable.
    """
    amount = _to_rupees(value)
    if amount is None:
        return "N/A"
    half = (amount + 1) // 2
    rounded = ((half + 500) // 1000) * 1000  # nearest thousand, half rounded up
    if rounded == 0:  # tiny amounts: don't round away to zero
        rounded = half
    return _format_indian_amount(rounded)


def _format_date(value) -> str:
    """Format a due date as e.g. '15 June'."""
    if _is_missing(value):
        return "N/A"
    if isinstance(value, datetime):
        return f"{value.day} {value.strftime('%B')}"
    try:
        dt = dateutil_parser.parse(str(value), dayfirst=True, fuzzy=True)
        return f"{dt.day} {dt.strftime('%B')}"
    except (ValueError, TypeError, OverflowError):
        return str(value).strip()


def _load_customer_context(path: Path) -> str:
    """Build CUSTOMER_CONTEXT from the second data row of the Excel sheet.

    Single streaming pass with openpyxl (read_only) instead of two full
    pandas parses -- keeps worker startup latency to a minimum.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    header_found = False
    columns: dict[str, int] = {}
    target_values: tuple | None = None
    data_rows_seen = 0

    for sheet_row in worksheet.iter_rows(values_only=True):
        if not header_found:
            labels = [str(v).strip().lower() for v in sheet_row if v is not None]
            if "customer name" in labels:
                header_found = True
                columns = {
                    str(v).strip().lower().replace(" ", "_"): idx
                    for idx, v in enumerate(sheet_row)
                    if v is not None
                }
            continue
        data_rows_seen += 1
        if data_rows_seen == 2:  # matches the previous df.iloc[1]
            target_values = sheet_row
            break

    workbook.close()

    if not header_found or target_values is None:
        raise RuntimeError(f"Could not locate customer row in {path}")

    def cell(col_name: str):
        idx = columns.get(col_name)
        return (
            target_values[idx] if idx is not None and idx < len(target_values) else None
        )

    name_parts = [
        str(cell(col)).strip()
        for col in ("customer_name", "last_name")
        if not _is_missing(cell(col))
    ]
    customer_name = " ".join(name_parts) or "Customer"

    # Remember the raw name so the PTP writer can find this exact row later.
    global LOADED_CUSTOMER_NAME
    LOADED_CUSTOMER_NAME = customer_name

    loan_amount = _format_indian_amount(cell("loan_amount"))
    due_amount = _format_indian_amount(cell("due_amount"))
    recommended_min = _recommended_min_payment(cell("due_amount"))
    due_date = _format_date(cell("due_date"))

    # Lender is fully data-driven: whatever the sheet says, with a neutral
    # fallback so the prompt never references a hardcoded bank.
    lender_value = cell("lender_name")
    lender_name = "N/A" if _is_missing(lender_value) else str(lender_value).strip()

    return textwrap.dedent(f"""
        Customer Name: {customer_name}
        Lender Name: {lender_name}
        Loan Amount: {loan_amount}
        Outstanding Amount: {due_amount}
        Recommended Minimum Payment (half of Outstanding, rounded to nearest 1000): {recommended_min}
        EMI Due Date: {due_date}
    """).strip()


CUSTOMER_CONTEXT = _load_customer_context(CUSTOMER_DATA_FILE)


# ---------------------------------------------------------------------------
# PTP normalization
#
# Excel must always store structured English data: PTP Amount as a plain number
# string ("5000") and PTP Date as DD-MM-YYYY ("25-06-2026"). The spoken
# conversation stays Hindi/Hinglish; only the values handed to
# record_promise_to_pay() and written to Excel are normalized here.
#
# Date policy: we always emit DD-MM-YYYY. When the source value has no year
# (e.g. "25 June", "पच्चीस जून"), we assume the current year. The LLM is
# instructed to resolve relative dates ("next Monday") to an absolute date with
# a year before calling the tool.
# ---------------------------------------------------------------------------

# Devanagari digits -> ASCII digits (२५ -> 25).
_DEV_DIGITS = str.maketrans("०१२३४५६७८९", "0123456789")

# Hindi number words. Includes the common 0-32 (covers day-of-month) plus the
# round tens, which is enough for the amounts and dates seen in this workflow.
_HINDI_UNITS = {
    "शून्य": 0,
    "एक": 1,
    "दो": 2,
    "तीन": 3,
    "चार": 4,
    "पाँच": 5,
    "पांच": 5,
    "छह": 6,
    "छः": 6,
    "छे": 6,
    "सात": 7,
    "आठ": 8,
    "नौ": 9,
    "दस": 10,
    "ग्यारह": 11,
    "बारह": 12,
    "तेरह": 13,
    "चौदह": 14,
    "पंद्रह": 15,
    "पन्द्रह": 15,
    "सोलह": 16,
    "सत्रह": 17,
    "अठारह": 18,
    "उन्नीस": 19,
    "बीस": 20,
    "इक्कीस": 21,
    "बाईस": 22,
    "तेईस": 23,
    "चौबीस": 24,
    "पच्चीस": 25,
    "पचीस": 25,
    "छब्बीस": 26,
    "सत्ताईस": 27,
    "अट्ठाईस": 28,
    "उनतीस": 29,
    "तीस": 30,
    "इकतीस": 31,
    "बत्तीस": 32,
    "चालीस": 40,
    "पचास": 50,
    "साठ": 60,
    "सत्तर": 70,
    "अस्सी": 80,
    "नब्बे": 90,
}
_HINDI_HUNDRED = {"सौ": 100}
_HINDI_BIG = {
    "हज़ार": 1_000,
    "हजार": 1_000,
    "लाख": 100_000,
    "करोड़": 10_000_000,
    "करोड": 10_000_000,
}
# Currency / filler words to ignore while parsing numbers.
_NUMBER_NOISE = {
    "रुपये",
    "रुपए",
    "रुपया",
    "रुपय",
    "रूपये",
    "रूपए",
    "रु",
    "रु.",
    "को",
    "और",
    "मात्र",
    "only",
    "rupaye",
    "rupees",
    "rs",
    "rs.",
}
_HINDI_MONTHS = {
    "जनवरी": "January",
    "फरवरी": "February",
    "फ़रवरी": "February",
    "मार्च": "March",
    "अप्रैल": "April",
    "मई": "May",
    "जून": "June",
    "जुलाई": "July",
    "अगस्त": "August",
    "सितंबर": "September",
    "सितम्बर": "September",
    "अक्टूबर": "October",
    "नवंबर": "November",
    "नवम्बर": "November",
    "दिसंबर": "December",
    "दिसम्बर": "December",
}

# Canonical Hindi word for each day-of-month (1-31), used to SPEAK a date aloud
# ("परसों" -> "एक जुलाई"). Reuses the 0-99 number words for a single source of truth.
_HINDI_DAY_WORDS = {n: _HINDI_NUMBERS_0_99[n] for n in range(1, 32)}

# English month name -> canonical Hindi month, derived from _HINDI_MONTHS.
_ENGLISH_TO_HINDI_MONTH = {english: hindi for hindi, english in _HINDI_MONTHS.items()}

# Hindi weekday names, indexed by datetime.weekday() (Monday == 0).
_HINDI_WEEKDAYS = (
    "सोमवार",
    "मंगलवार",
    "बुधवार",
    "गुरुवार",
    "शुक्रवार",
    "शनिवार",
    "रविवार",
)


def _spoken_hindi_date(dt: datetime) -> str:
    """Render a date as naturally spoken Hindi words ('1 July' -> 'एक जुलाई')."""
    day = _HINDI_DAY_WORDS.get(dt.day, str(dt.day))
    month = _ENGLISH_TO_HINDI_MONTH.get(dt.strftime("%B"), dt.strftime("%B"))
    return f"{day} {month}"


def _relative_date_reference(today: datetime | None = None) -> str:
    """Build a concrete relative-date table for the prompt.

    Resolving "कल"/"next Monday" to a real date is date arithmetic, which LLMs
    do unreliably. We compute the dates here so the model only has to look them
    up: it speaks the Hindi form shown and passes the DD-MM-YYYY shown to the
    tool, with no arithmetic of its own.
    """
    today = today or datetime.now()

    def line(label: str, dt: datetime) -> str:
        weekday = _HINDI_WEEKDAYS[dt.weekday()]
        return f'- "{label}" = {_spoken_hindi_date(dt)} ({dt.strftime("%d-%m-%Y")}, {weekday})'

    lines = [
        line("आज / today", today),
        line("कल / tomorrow", today + timedelta(days=1)),
        line("परसों / day after tomorrow", today + timedelta(days=2)),
    ]

    # "N दिन बाद" / "in N days" for the next 15 days, so any near "X days later"
    # is a direct lookup — this is exactly where the model used to hallucinate
    # dates. Capped at 15 (the realistic phone-call range) to keep the prompt
    # lean; beyond that customers give an explicit date, and the prompt's
    # "date not in the table" fallback covers it. Bump this back up if you see
    # customers naming day-counts past two weeks.
    lines.append("")
    lines.append('Days from today ("N दिन बाद" / "in N days"):')
    for n in range(1, 16):
        dt = today + timedelta(days=n)
        weekday = _HINDI_WEEKDAYS[dt.weekday()]
        lines.append(
            f'- "{_HINDI_NUMBERS_0_99[n]} दिन बाद / in {n} day(s)" = '
            f"{_spoken_hindi_date(dt)} ({dt.strftime('%d-%m-%Y')}, {weekday})"
        )

    # Next occurrence of each weekday (strictly after today, within 1-7 days),
    # so phrases like "अगले सोमवार" / "next Monday" resolve correctly.
    lines.append("")
    lines.append('Next weekday ("अगले <din>" / "next <weekday>"):')
    for offset in range(1, 8):
        dt = today + timedelta(days=offset)
        weekday = _HINDI_WEEKDAYS[dt.weekday()]
        lines.append(
            f'- "अगले {weekday} / next {dt.strftime("%A")}" = '
            f"{_spoken_hindi_date(dt)} ({dt.strftime('%d-%m-%Y')})"
        )

    # Month-scale horizon: customers often say "एक महीने बाद" / "अगले महीने"
    # without a day. We estimate ~30 days out so the agent can propose a
    # concrete date instead of stalling. Kept as a single row to stay lean.
    lines.append("")
    lines.append('Longer horizons ("एक महीने बाद" / "अगले महीने" / "in a month"):')
    dt = today + timedelta(days=30)
    weekday = _HINDI_WEEKDAYS[dt.weekday()]
    lines.append(
        f'- "एक महीने बाद / अगले महीने / in about a month (~30 days)" = '
        f"{_spoken_hindi_date(dt)} ({dt.strftime('%d-%m-%Y')}, {weekday})"
    )
    return "\n".join(lines)


def _hindi_words_to_number(text: str) -> int | None:
    """Parse Hindi number words into an integer ('पाँच हज़ार' -> 5000).

    Returns None if any token is not a recognized number word.
    """
    result = current = 0
    found = False
    for token in str(text).replace(",", " ").split():
        tok = token.strip()
        if not tok or tok in _NUMBER_NOISE:
            continue
        if tok in _HINDI_UNITS:
            current += _HINDI_UNITS[tok]
        elif tok in _HINDI_HUNDRED:
            current = (current or 1) * _HINDI_HUNDRED[tok]
        elif tok in _HINDI_BIG:
            result += (current or 1) * _HINDI_BIG[tok]
            current = 0
        else:
            return None
        found = True
    return result + current if found else None


def normalize_ptp_amount(value) -> str | None:
    """Normalize an amount to a plain integer string ('चार हज़ार रुपए' -> '4000').

    Handles ASCII digits, Devanagari digits, and Hindi number words. Returns
    None when nothing numeric can be extracted.
    """
    if _is_missing(value):
        return None
    text = str(value).strip().replace("₹", "").replace(",", "").translate(_DEV_DIGITS)
    # Already a plain number, e.g. "5000" or "5000.0".
    if re.fullmatch(r"\d+(?:\.\d+)?", text.strip()):
        return str(int(float(text.strip())))
    # Hindi words, e.g. "नौ हज़ार पाँच सौ".
    number = _hindi_words_to_number(text)
    if number is not None:
        return str(number)
    # Last resort: pull any embedded digits ("₹5,000 only" -> "5000").
    digits = "".join(re.findall(r"\d+", text))
    return digits or None


def normalize_ptp_date(value, today: datetime | None = None) -> str | None:
    """Normalize a date to DD-MM-YYYY ('दस जुलाई' -> '10-07-2026').

    Handles existing DD-MM-YYYY input, Devanagari digits, Hindi month names,
    and Hindi day-number words. Missing years default to the current year.
    Returns None when the value cannot be parsed.
    """
    if _is_missing(value):
        return None
    today = today or datetime.now()
    text = str(value).strip().translate(_DEV_DIGITS)

    # Fast path: already a recognizable explicit format.
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
        except ValueError:
            pass

    # Translate Hindi months and day-number words into English/digits.
    for hindi, english in _HINDI_MONTHS.items():
        text = text.replace(hindi, f" {english} ")
    text = " ".join(
        str(_HINDI_UNITS[t]) if t in _HINDI_UNITS else t for t in text.split()
    )

    try:
        # default supplies the year (and any missing parts) for inputs like
        # "10 July"; fuzzy lets us ignore leftover particles such as "को".
        dt = dateutil_parser.parse(text, default=today, dayfirst=True, fuzzy=True)
        return dt.strftime("%d-%m-%Y")
    except (ValueError, OverflowError):
        return None


def migrate_ptp_columns(path: Path = CUSTOMER_DATA_FILE) -> int:
    """Rewrite existing PTP Amount / PTP Date cells into structured English.

    Only the two PTP columns are touched; all other columns and the sheet
    layout are preserved. Returns the number of cells changed.
    """
    try:
        workbook = load_workbook(path)
    except Exception:
        logger.exception("Could not open %s for PTP migration", path)
        return 0

    worksheet = workbook.active

    header_row_idx = None
    columns: dict[str, int] = {}
    for r_idx, sheet_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        labels = {
            str(v).strip().lower(): c
            for c, v in enumerate(sheet_row, start=1)
            if v is not None
        }
        if "customer name" in labels:
            header_row_idx = r_idx
            columns = labels
            break

    amount_col = columns.get("ptp amount")
    date_col = columns.get("ptp date")
    if header_row_idx is None or not (amount_col and date_col):
        logger.error("PTP columns not found in %s; nothing migrated", path)
        return 0

    changed = 0
    for r_idx in range(header_row_idx + 1, worksheet.max_row + 1):
        amount_cell = worksheet.cell(row=r_idx, column=amount_col)
        if not _is_missing(amount_cell.value):
            new_amount = normalize_ptp_amount(amount_cell.value)
            if new_amount is not None and str(amount_cell.value) != new_amount:
                logger.info(
                    "Row %d PTP amount: %r -> %r", r_idx, amount_cell.value, new_amount
                )
                amount_cell.value = new_amount
                changed += 1

        date_cell = worksheet.cell(row=r_idx, column=date_col)
        if not _is_missing(date_cell.value):
            new_date = normalize_ptp_date(date_cell.value)
            if new_date is not None and str(date_cell.value) != new_date:
                logger.info(
                    "Row %d PTP date: %r -> %r", r_idx, date_cell.value, new_date
                )
                date_cell.value = new_date
                changed += 1

    if changed:
        workbook.save(path)
    logger.info("PTP migration complete: %d cell(s) updated in %s", changed, path)
    return changed


def save_promise_to_pay(
    customer_name: str,
    ptp_amount: str,
    ptp_date: str,
    path: Path = CUSTOMER_DATA_FILE,
) -> bool:
    """Write PTP Amount and PTP Date for a customer back into the Excel file.

    Only the two PTP cells of the matching row are touched; every other value
    and the sheet layout are left untouched. Returns True on success.
    """
    # Final safety gate: never persist a partial promise. The tool layer also
    # checks this, but we re-check here so the file can never hold half a PTP.
    if not str(ptp_amount).strip() or not str(ptp_date).strip():
        logger.warning(
            "Refusing to save incomplete PTP (amount=%r, date=%r)", ptp_amount, ptp_date
        )
        return False

    try:
        workbook = load_workbook(path)
        worksheet = workbook.active

        # Locate the header row and the columns we care about, by name, so the
        # code is resilient to column reordering or a title banner above them.
        header_row_idx = None
        columns: dict[str, int] = {}
        for r_idx, sheet_row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            labels = {
                str(v).strip().lower(): c
                for c, v in enumerate(sheet_row, start=1)
                if v is not None
            }
            if "customer name" in labels:
                header_row_idx = r_idx
                columns = labels
                break

        if header_row_idx is None:
            logger.error("No header row with 'Customer Name' found in %s", path)
            return False

        name_col = columns.get("customer name")
        amount_col = columns.get("ptp amount")
        date_col = columns.get("ptp date")
        if not (name_col and amount_col and date_col):
            logger.error("Missing PTP columns in %s (found: %s)", path, list(columns))
            return False

        # Find the row for the loaded customer.
        target_row = None
        for r_idx in range(header_row_idx + 1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=r_idx, column=name_col).value
            if (
                cell_value is not None
                and str(cell_value).strip().lower() == customer_name.strip().lower()
            ):
                target_row = r_idx
                break

        if target_row is None:
            logger.error("Customer %r not found in %s", customer_name, path)
            return False

        worksheet.cell(row=target_row, column=amount_col, value=ptp_amount)
        worksheet.cell(row=target_row, column=date_col, value=ptp_date)
        workbook.save(path)
        logger.info(
            "Saved PTP for %s (row %d): amount=%r date=%r",
            customer_name,
            target_row,
            ptp_amount,
            ptp_date,
        )
        return True
    except Exception:
        logger.exception("Failed to save PTP to %s", path)
        return False


# ---------------------------------------------------------------------------
# Call outcome
#
# Exactly one outcome is recorded per call, written once near the end. If the
# conversation changes direction, the most recent valid outcome wins (the cell
# is simply overwritten). Outcome lives in its own column and its own tool, so
# it never interferes with the PTP Amount / PTP Date cells.
# ---------------------------------------------------------------------------

# The only outcomes the bot may store. Canonical spelling -> lookup key.
VALID_OUTCOMES = [
    "Abusive Customer",
    "Wrong Person",
    "Dont Call Me",
    "Refuse to pay",
    "Already Paid",
    "Call Back",
    "Will Pay",
    "Paying Rightaway",
    "Connect me to Humans",
    "Call Disconnected - No outcome",
    "Discussed - No Outcome",
]
_OUTCOME_LOOKUP = {o.lower(): o for o in VALID_OUTCOMES}


def normalize_outcome(value) -> str | None:
    """Map a model-supplied outcome to its canonical spelling.

    Case-insensitive; returns None for anything not in VALID_OUTCOMES so an
    invented outcome can never be written.
    """
    if _is_missing(value):
        return None
    return _OUTCOME_LOOKUP.get(str(value).strip().lower())


def save_call_outcome(
    customer_name: str,
    outcome: str,
    path: Path = CUSTOMER_DATA_FILE,
) -> bool:
    """Write the final call Outcome for a customer back into the Excel file.

    Only the single Outcome cell of the matching row is touched. Intentionally
    last-write-wins: the final outcome of the call overwrites any earlier one.
    Returns True on success.
    """
    canonical = normalize_outcome(outcome)
    if canonical is None:
        logger.warning("Refusing to save invalid outcome %r", outcome)
        return False

    try:
        workbook = load_workbook(path)
        worksheet = workbook.active

        header_row_idx = None
        columns: dict[str, int] = {}
        for r_idx, sheet_row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            labels = {
                str(v).strip().lower(): c
                for c, v in enumerate(sheet_row, start=1)
                if v is not None
            }
            if "customer name" in labels:
                header_row_idx = r_idx
                columns = labels
                break

        if header_row_idx is None:
            logger.error("No header row with 'Customer Name' found in %s", path)
            return False

        name_col = columns.get("customer name")
        outcome_col = columns.get("outcome")
        if not (name_col and outcome_col):
            logger.error(
                "Missing Outcome column in %s (found: %s)", path, list(columns)
            )
            return False

        target_row = None
        for r_idx in range(header_row_idx + 1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=r_idx, column=name_col).value
            if (
                cell_value is not None
                and str(cell_value).strip().lower() == customer_name.strip().lower()
            ):
                target_row = r_idx
                break

        if target_row is None:
            logger.error("Customer %r not found in %s", customer_name, path)
            return False

        worksheet.cell(row=target_row, column=outcome_col, value=canonical)
        workbook.save(path)
        logger.info(
            "Saved outcome for %s (row %d): %r", customer_name, target_row, canonical
        )
        return True
    except Exception:
        logger.exception("Failed to save outcome to %s", path)
        return False


# ---------------------------------------------------------------------------
# Agent-initiated hangup
#
# Deleting the room ends the call for both web and SIP/telephony participants,
# and fires the session "close" event that runs outcome classification. Kept as
# a module-level function so the end_call tool stays trivially unit-testable
# (tests monkeypatch _hangup instead of touching the LiveKit API).
# ---------------------------------------------------------------------------
# Closing/farewell markers used to detect when Priya has ended the call. These
# are deliberately multi-word, closing-only phrases: a bare "धन्यवाद" also
# appears in the OPENING greeting, so matching on that alone would cut the call
# right after it starts. Romanized variants are included as a fallback.
_FAREWELL_MARKERS = (
    "आपका दिन शुभ",
    "दिन शुभ हो",
    "शुभ दिन",
    "अलविदा",
    "धन्यवाद आपका समय",
    "aapka din shubh",
    "din shubh ho",
    "shubh din",
    "alvida",
)


def _is_farewell(text) -> bool:
    """True if the agent's line is a closing farewell (so the call can end)."""
    if not text:
        return False
    # Strip punctuation (danda "।", commas, etc.) but keep the full Devanagari
    # block U+0900-U+097F -- \w alone drops combining vowel marks (matras),
    # which would mangle the Hindi words. Then collapse whitespace.
    normalized = re.sub(r"[^\w\sऀ-ॿ]", " ", str(text))
    normalized = re.sub(r"\s+", " ", normalized).strip().lower()
    return any(marker in normalized for marker in _FAREWELL_MARKERS)


async def _hangup() -> None:
    """End the call by deleting the room."""
    job_ctx = get_job_context()
    if job_ctx is None:
        logger.warning("Hangup requested but no job context is available")
        return
    try:
        await job_ctx.api.room.delete_room(
            api.DeleteRoomRequest(room=job_ctx.room.name)
        )
        logger.info(">>> HANGUP room=%r deleted", job_ctx.room.name)
    except Exception:
        logger.exception("Failed to delete room during hangup")


# ---------------------------------------------------------------------------
# Post-call outcome classification
#
# The outcome is NOT recorded by the voice agent via a tool. Asking the speaking
# model to emit a tool call inside its closing turn caused the tool-call syntax
# to leak into the customer transcript (and the call was never dispatched).
#
# Instead, once the call ends (including on disconnect), we run a single,
# non-voice classification over the transcript and persist exactly one outcome.
# This keeps outcome selection completely off the voice path, so nothing
# internal can ever reach the customer.
# ---------------------------------------------------------------------------

_OUTCOME_DEFINITIONS = """\
- "Abusive Customer" — the customer is abusive, threatening, or uses offensive language.
- "Wrong Person" — the person reached is not the customer (wrong number, or no such person / says it is not them).
- "Dont Call Me" — the customer asks not to be called again.
- "Refuse to pay" — the customer clearly refuses to pay and makes no commitment.
- "Already Paid" — the customer says they have already paid, or that the account is settled.
- "Call Back" — the customer (or whoever answered) asks to be called back later, or says the customer is busy / not available right now.
- "Will Pay" — the customer commits to pay on a future date.
- "Paying Rightaway" — the customer says they are paying now / immediately.
- "Connect me to Humans" — the customer asks to be transferred to a human agent.
- "Call Disconnected - No outcome" — the call dropped before any resolution was reached.
- "Discussed - No Outcome" — the matter was discussed but none of the above applies.
"""


def _match_outcome_in_text(text) -> str | None:
    """Map the classifier's free-form reply to one canonical outcome label."""
    if _is_missing(text):
        return None
    cleaned = str(text).strip().strip("\"'`.").strip()
    exact = normalize_outcome(cleaned)
    if exact is not None:
        return exact
    # The model may wrap the label in a sentence; find the first label present.
    lowered = str(text).lower()
    for outcome in VALID_OUTCOMES:
        if outcome.lower() in lowered:
            return outcome
    return None


def _transcript_from_history(history: ChatContext) -> str:
    """Render the call's user/assistant turns into a plain transcript."""
    lines: list[str] = []
    for item in history.items:
        role = getattr(item, "role", None)
        if role not in ("user", "assistant"):
            continue
        text = (item.text_content or "").strip()
        if not text:
            continue
        lines.append(f"{'Customer' if role == 'user' else 'Agent'}: {text}")
    return "\n".join(lines)


def _is_valid_hindi_text(text: str) -> bool:
    """Check if transcribed text looks like Hindi/Hinglish, not garbage.

    Garbage indicators: text in Spanish, French, random Latin scripts that
    aren't English + Hindi. Returns False if it looks like a bad STT output.
    """
    if not text or len(text.strip()) < 2:
        return True  # too short to judge, allow it

    # Count scripts
    devanagari_chars = sum(1 for c in text if "\u0900" <= c <= "\u097f")
    latin_chars = sum(1 for c in text if c.isascii() and c.isalpha())
    total_alpha = devanagari_chars + latin_chars

    if total_alpha == 0:
        return True  # no letters, probably just punctuation/numbers

    # Hindi/Hinglish should have at least 20% Devanagari OR be mostly English
    # If it's a mix with very little Devanagari, it's likely garbage (Spanish etc)
    devanagari_ratio = devanagari_chars / total_alpha
    latin_ratio = latin_chars / total_alpha

    if total_alpha > 5 and devanagari_ratio < 0.15 and latin_ratio < 0.85:
        # Lots of text, almost no Devanagari, and not pure English either
        logger.warning(
            "Garbage STT detected: %r (%.1f%% Devanagari, %.1f%% Latin)",
            text,
            devanagari_ratio * 100,
            latin_ratio * 100,
        )
        return False

    return True


# ---------------------------------------------------------------------------
# STT noise filtering (filler sounds + duplicate finals)
#
# These are applied in Assistant.stt_node, which is the only reliable place to
# drop an utterance BEFORE it becomes a user turn: there is no "user_speech"
# session event, so an .on(...) handler can never block one. Dropping the
# FINAL_TRANSCRIPT here means the utterance never reaches the transcript, the
# LLM, or preemptive generation.
# ---------------------------------------------------------------------------

# Filler sounds with any elongation: "hmm", "uhh", "uhmm", "ahh", "err", ...
_FILLER_TOKEN_RE = re.compile(
    r"u+h+m*|u+m+|u+h*|h+m+|h+|m+|a+h+|e+r+|e+h+|o+h+|huh|hmm+"
)
# Devanagari spellings the STT may emit for the same sounds. Deliberately
# excludes meaningful words like "हाँ" (yes) / "ना" (no).
_DEVANAGARI_FILLERS = {
    "हम्म",
    "हम्म्म",
    "हँ",
    "अं",
    "अँ",
    "उम",
    "उह",
    "अह",
    "आह",
    "ऊँ",
    "एं",
}
# Window within which an identical final transcript is treated as an STT
# duplicate rather than the customer genuinely repeating themselves.
_STT_DEDUP_WINDOW_S = 6.0


def _is_filler_only(text: str) -> bool:
    """True if the utterance is nothing but filler sounds ('hmm', 'उह')."""
    if _is_missing(text):
        return False
    tokens = re.findall(r"[a-zऀ-ॿ]+", str(text).lower())
    if not tokens:
        return False
    for tok in tokens:
        if tok in _DEVANAGARI_FILLERS or _FILLER_TOKEN_RE.fullmatch(tok):
            continue
        return False
    return True


def _dedup_key(text: str) -> str:
    """Normalize an utterance (lowercase, punctuation-stripped) for dedup compare."""
    cleaned = re.sub(r"[^\w]+", " ", str(text).lower())
    return re.sub(r"\s+", " ", cleaned).strip()


async def classify_and_save_outcome(history: ChatContext, customer_name: str) -> None:
    """Classify the finished call into exactly one outcome and persist it."""
    try:
        transcript = _transcript_from_history(history)
        if not transcript:
            # Nothing was exchanged — the call dropped before any resolution.
            save_call_outcome(customer_name, "Call Disconnected - No outcome")
            return

        ctx = ChatContext.empty()
        ctx.add_message(
            role="system",
            content=(
                "You classify the outcome of a debt-collection phone call. "
                "Read the transcript and pick the single label that best describes "
                "how the call ended. If the customer changed direction during the "
                "call, use the most recent valid outcome. Reply with the label "
                "text ONLY, nothing else.\n\n" + _OUTCOME_DEFINITIONS
            ),
        )
        ctx.add_message(role="user", content=f"Transcript:\n{transcript}\n\nOutcome:")

        # A small, fast model — this is a trivial one-label classification, so we
        # avoid the heavy voice model. temperature=0 makes it deterministic and
        # max_completion_tokens caps it to just the label, cutting latency to ~1s.
        classifier = inference.LLM(
            model="openai/gpt-4o-mini",
            extra_kwargs={"temperature": 0, "max_completion_tokens": 16},
        )
        raw = ""
        try:
            async with classifier.chat(chat_ctx=ctx) as stream:
                async for chunk in stream:
                    if chunk.delta and chunk.delta.content:
                        raw += chunk.delta.content
        finally:
            await classifier.aclose()

        outcome = _match_outcome_in_text(raw) or "Discussed - No Outcome"
        logger.info(
            ">>> OUTCOME-CLASSIFY customer=%r raw=%r -> %r",
            customer_name,
            raw.strip(),
            outcome,
        )
        save_call_outcome(customer_name, outcome)
    except Exception:
        logger.exception("Outcome classification failed for %r", customer_name)


# NOTE: The strict identity-verification flow was disabled in favor of a simple
# name confirmation (see the "IDENTITY CHECK (SIMPLE)" block in the prompt).
# The previous behavior is kept here, commented out, so it can be restored:
#
#   IDENTITY VERIFICATION
#   Before identity verification:
#     * Do not reveal loan amount / due amount / due date / account status.
#     * Do not reveal any customer-specific information.
#   If identity cannot be verified, politely end the call.
def _build_agent_llm():
    """Build the conversation LLM.

    If AWS credentials are present, use Claude Haiku on Amazon Bedrock in
    ap-south-1 (Mumbai) — co-located compute for low latency to Indian callers.
    This is the co-location latency test. Otherwise fall back to OpenRouter
    gpt-4.1-mini (latency-sorted routing) so dev/tests work without AWS. Toggle
    by setting or unsetting AWS_ACCESS_KEY_ID in .env.local.

    BEDROCK_MODEL must be an inference-profile id available in the target
    region. For ap-south-1 (Mumbai) the latest Claude is served only via the
    Global cross-region inference profile (the "global." prefix), not the bare
    model id — see AWS's "Claude models in India" guidance.
    """
    if os.environ.get("AWS_ACCESS_KEY_ID"):
        return aws.LLM(
            model=os.environ.get(
                "BEDROCK_MODEL", "global.anthropic.claude-haiku-4-5-20251001-v1:0"
            ),
            region=os.environ.get("AWS_REGION", "ap-south-1"),
            temperature=0,
        )
    return openai.LLM.with_openrouter(
        model="openai/gpt-4.1-mini",
        api_key=os.environ.get("OPEN_ROUTER_KEY"),
        provider={"sort": "latency"},
    )


class Assistant(Agent):
    def __init__(self) -> None:
        # Resolve relative dates ("कल", "next Monday") in code and hand the model
        # a concrete lookup table, so it never has to do date arithmetic itself.
        now = datetime.now()
        relative_dates = _relative_date_reference(now)
        super().__init__(
            llm=_build_agent_llm(),
            instructions=textwrap.dedent(f"""
You are Priya, an experienced debt resolution specialist calling on behalf of a financial services company.

Your customer's details (name, lender, dues) are in the CUSTOMER INFORMATION section at the END of these instructions. Follow it together with the IDENTITY CHECK there before revealing any account details.

================================================
LANGUAGE
================================================



* Default: Hinglish — Hindi words in Devanagari, BFSI/business terms in English. Do not ask language preference; switch permanently to English only when explicitly requested.
* Hindi customer: respond primarily in Devanagari Hindi with natural English BFSI terms.
* English customer: respond entirely in English.
* Hinglish customer: natural Indian Hinglish — Hindi in Devanagari, business terms in English.

ENGLISH SWITCH (HIGH PRIORITY — overrides the default language rules above):
Default language is Hinglish. NEVER start speaking English on your own — stay in Hindi/Hinglish for the entire call. ONLY switch to English if the customer EXPLICITLY asks you to (e.g. "speak in English"). If they do, treat it as a PERMANENT instruction for the remainder of the conversation. After switching:
- NEVER switch back to Hindi or Hinglish on your own.
- NEVER mix English with Hindi.
- Continue responding ONLY in English until the customer explicitly asks to switch back.

Correct: "नमस्ते राहुल जी।" · "आपके personal loan account में 8000 ki payment due है।" · "क्या आप payment आज कर पाएंगे?" · "मैं callback request बना देती हूँ।" · "क्या आप human agent से बात करना चाहेंगे?"
Avoid overly formal Hindi. Incorrect: "मैं पुनर्भुगतान संग्रहण विभाग से बोल रही हूँ।" · "आपका ऋण खाता बकाया है।"

================================================
BFSI TERMINOLOGY (IMPORTANT)
================================================

Always use the standard English BFSI term, written in Latin script, for the
words below — prefer the English term over any pure-Hindi word in EVERY sentence,
including greetings, standard lines, and confirmations. NEVER use their
pure/literary Hindi translations.

* "payment" — never भुगतान
* "amount" — never राशि (do not use राशि at all; just say "payment")
* "outstanding" / "due" / "overdue" — never बकाया
* "loan" — never ऋण
* "EMI" / "installment" — never किस्त / क़िस्त
* "due date" — never देय तिथि / नियत तिथि
* "confirm" — never पुष्टि
* "verify" — never सत्यापन
* "review" — never समीक्षा
* "settlement" — never समझौता
* "account" — never लेखा

When asking how much the customer will pay, ask "कितनी payment कर पाएंगे?" —
never insert राशि (do not say "कितनी राशि payment कर पाएंगे").

Correct: "आप कितनी payment कर पाएंगे?" / "आपका payment अभी outstanding है।"
Incorrect: "आप कितनी राशि payment कर पाएंगे?" / "आपकी राशि अभी बकाया है।"

* Start every call in natural Hindi/Hinglish; if unclear, continue in Hindi/Hinglish. Never write Hindi in Roman script. Speak only Hindi, Hinglish, or English — never Spanish, French, German, or any other language.

================================================
HINDI STYLE (FOR SARVAM TTS)
============================

Optimize all Hindi for natural Sarvam TTS: short, everyday spoken Hindi (not formal/literary), correct Devanagari spelling and matras, short sentences that sound natural on a phone call.

================================================
SPEECH FORMAT
=============

When speaking to customers:

* Natural conversational language; one question at a time; pause and wait for the response.
* Never use digits for money or dates — speak amounts and dates in Hindi words when using Hindi. Read OTPs and reference numbers digit by digit.

Examples: ₹18,750 → अठारह हज़ार सात सौ पचास रुपये · ₹2,00,000 → दो लाख रुपये · 15 June → पंद्रह जून

================================================
TOOL FORMAT
===========

When calling tools:

* Amount must be numeric only.
* Date must be DD-MM-YYYY.
* Never pass Hindi text to tools.
* Spoken language and tool arguments are separate concerns.

Example — Customer: "मैं पाँच हज़ार रुपये पच्चीस जून को दे दूँगा" → Tool: amount="5000", date="25-06-2026".

Today's date is {now.strftime("%d-%m-%Y")} ({_HINDI_WEEKDAYS[now.weekday()]}).

RELATIVE DATES — use this exact table; you are FORBIDDEN from calculating any date
yourself (that has produced wrong dates). Find the matching row, SPEAK the Hindi form
shown, and pass its DD-MM-YYYY to the tool. This includes every "N दिन बाद" / "N days
later" row (e.g. "दस दिन बाद", "तीन दिन बाद") — never add the days yourself.

{relative_dates}

When confirming a relative date with the customer, always state the date from the
matching row (e.g. "दस दिन बाद से मतलब <table date>") — never a date you computed.
For "अगले हफ्ते" / "next week", confirm a specific day from the table first. For a
date not in the table (for example "महीने की 5 तारीख" / "5th of the month"), use
the explicit day with the current or next month as the customer means, in
DD-MM-YYYY.

For a vague month-scale timeframe ("एक महीने बाद", "अगले महीने", "one month"), do
NOT get stumped: use the ~30-day "Longer horizons" row above as a concrete
estimate, propose that exact date, and ask if the customer can pay by then
(e.g. "जी, तो क्या आप <table date> तक payment कर पाएंगे?"). Record the Promise To
Pay with that date once they confirm.

================================================
INDIAN DATE INTERPRETATION
================================================

Customers may express payment dates informally.

Examples: "कल" · "परसों" · "अगले सोमवार" · "अगले हफ्ते" · "महीने की 5 तारीख" · "salary आते ही" — interpret these naturally in conversation.

Before recording a Promise To Pay, always confirm the exact payment date if there is any ambiguity.

Example (uses the table above — "कल" resolves to {_spoken_hindi_date(now + timedelta(days=1))}):

Customer:
"कल कर दूँगा।"

Agent:
"जी, confirm कर दूँ — क्या आप {_spoken_hindi_date(now + timedelta(days=1))} को payment करेंगे?"

Only record a Promise To Pay after the customer confirms the specific date.

================================================
TOOL FAILURE HANDLING
=====================

If a tool fails: do not tell the customer, continue normally, acknowledge their commitment, and close professionally.

================================================
VOICE STYLE
===========

You are on a live phone call. Be warm, friendly, and conversational — like a helpful, empathetic person, not a chatbot or support script. Use the customer's name naturally and sound relaxed, calm, and professional (there is a pending payment) — never robotic or scripted. Use natural Hinglish, not pure Hindi words like pushti or samjhauta. Natural fillers ("जी", "समझ गई", "ठीक है", "बिलकुल") are fine but don't overuse them.

BREVITY (critical — you have been far too long and repetitive): Keep EVERY reply to ONE or TWO short spoken sentences — concise and to the point, but natural (not clipped or terse). Lead with a brief acknowledgment ("जी", "ठीक है", "समझ गई") so speech starts right away, then say the one thing that matters. Ask only ONE question, and ask it exactly ONCE — NEVER repeat, rephrase, or restate the same question or point within a single reply, and never pad with extra explanation. Never produce more than two sentences. If you notice yourself restating something, stop immediately.

================================================
CONVERSATION MEMORY
===================

Maintain conversation state throughout the call. Track internally what the
customer has already provided — identity confirmation, correct-person
confirmation, payment amount, payment date, inability-to-pay reason, dispute
reason, and any callback, settlement, or escalation request — and treat it as
known for the rest of the call. Never re-ask for information already collected
unless clarification is genuinely required. Always move the conversation forward.

================================================
ANTI-REPETITION
===============

This is extremely important. Never repeat introductions, verification requests,
account or payment explanations, escalation or callback messages, or previously
answered questions. If the customer already answered, acknowledge it, build on
it, and move forward — never re-ask the same thing using different wording.

Bad: agent asks "When can you make the payment?", customer says "Next Monday", agent then asks "What date can you make the payment?" (wrong — already answered). Good: "Understood. How much will you be able to pay on Monday?"

If the customer repeats the same answer, do not re-ask — gather the next required
information, offer an alternative, move toward resolution, or escalate.

Every response must do at least one of: collect new information, confirm a
commitment, resolve an objection, or progress toward closure or escalation. If no
new information emerges after two exchanges on the same topic, move to the next
step or escalate. Avoid conversational loops at all costs.

================================================
INTERRUPTION HANDLING
=====================

If the customer interrupts: address what they said immediately and naturally. NEVER restart the call or return to the opening intro. If you were cut off while stating the outstanding amount or EMI due date (or another important detail), after answering the interruption, RESUME and finish delivering that exact information — continue from where you were cut off; never drop the figure and never start over from the greeting.

================================================
CALL FLOW
=========

1. Introduce yourself as Priya.
2. Confirm the correct person.
3. Verify identity.
4. Discuss the account.
5. Work toward resolution.
6. Close or escalate.

This flow is guidance, not a script — sound natural.

================================================
PAYMENT RESOLUTION PRINCIPLES
=============================

When discussing repayment:

* Understand the customer's intent first, then work toward a specific commitment — prefer specific dates over vague promises and specific amounts over general willingness to pay.

* If the customer asks how much they should pay (e.g. "कितना दूँ?", "how much should I pay?"), recommend the "Recommended Minimum Payment" shown in CUSTOMER INFORMATION (half the Outstanding Amount, rounded to the nearest thousand) — speak the exact amount shown, in Hindi words. Encourage the full Outstanding if they can manage it. (If it shows N/A, recommend roughly half the Outstanding Amount.)

Tentative statements ("Maybe", "I'll try", "Let's see", "Hopefully", "Should be able to") are NOT commitments — do not treat them as confirmed promises.

================================================
OUTCOMES
========

PROMISE TO PAY — the moment the customer states an amount, treat it as SET and move on. Do NOT re-ask it or keep re-confirming it — repeating the amount is annoying. Only change it if the customer themselves changes it. Collect the payment date the same way. Confirm the amount and date together at most ONCE, right before recording — then call record_promise_to_pay and close politely. Keep speaking naturally in Hindi/Hinglish, but pass structured English values to the tool (amount: digits only in rupees; date: DD-MM-YYYY). Do not invent values; use only what the customer explicitly provided.

UNABLE TO PAY — explore one realistic future payment date.

REFUSAL TO PAY — ask the reason once, attempt one resolution, escalate if unresolved.

ALREADY PAID — collect payment date and reference number, mark for review.

DISPUTE — collect the dispute reason, escalate for review.

CALLBACK — collect callback date and time, confirm once.

SETTLEMENT REQUEST — collect request details, escalate.

WRONG PARTY — reveal no account information; apologize briefly and close with the farewell so the call ends.

================================================
CALL CLOSING
============

Once a clear outcome has been reached: do not reopen negotiation or ask new questions; summarize the next action briefly and end politely, in under two sentences.

When the conversation is finished (an outcome has been reached, or it is a wrong
party, or there is nothing more to do), end with a short farewell. Your final
sentence MUST be a clear goodbye, ending with "धन्यवाद, आपका दिन शुभ हो।". The
call ends automatically right after you say it, so say it only when you are truly
done — do not say it mid-conversation.

NEVER speak, type, read aloud, or output any function name, code, JSON, or tool
syntax (for example never produce "end_call", "functions.end_call", or
"{{ outcome_reached: ... }}"). Speak only natural Hindi to the customer. The call
ending is handled automatically from your spoken farewell — you do not call any
tool to end it.

================================================
ESCALATION
==========

Escalate when: repeated refusal, supervisor requested, settlement requested, legal threats mentioned, or dispute requires review.

When escalating, say:

"मैं आपका केस सीनियर टीम को रिव्यू के लिए फॉरवर्ड कर रही हूँ।"

Do not continue negotiation afterward.

================================================
IMPORTANT
=========

Never mention being AI. Never use markdown or emojis in speech.

Never congratulate the customer or use celebratory words such as "बधाई", "मुबारक", or "congratulations" — a payment is pending, there is nothing to celebrate. Open with a simple "नमस्ते" and get to the point directly, warmly, and politely.

Never speak the ₹ symbol or any digits for money. Once identity is confirmed, PROACTIVELY tell the customer their EXACT Outstanding Amount — say the specific figure from CUSTOMER INFORMATION in Hindi words (for example "आपका outstanding नौ हज़ार पाँच सौ रुपये है", never "₹9500" or "9500"). NEVER be vague about the figure: never say "कुछ outstanding", "कुछ राशि", "थोड़ा payment", "some amount", or any placeholder in place of the real number. Stating a non-specific amount is only allowed in the pre-confirmation opening line.

Likewise, state the EXACT EMI Due Date from CUSTOMER INFORMATION — the specific date in Hindi words (for example "आपकी EMI की due date पंद्रह जून थी") — never just say the payment is "due" or "pending" without naming the date.

Stay calm, human, and conversational.

Your goal is to achieve a clear resolution and collect the next best action.

================================================
CUSTOMER INFORMATION
====================

{CUSTOMER_CONTEXT}

Use this information naturally once the person confirms they are {LOADED_CUSTOMER_NAME}. Refer to the lender by the Lender Name on file (for example, "आपके HDFC Bank लोन को लेकर") only after they confirm, and never assume a lender not listed above. Do not reveal loan amount, due amount, due date, or any account details until the person confirms they are {LOADED_CUSTOMER_NAME}.

================================================
IDENTITY CHECK (SIMPLE)
=======================

Open by confirming, by name, that you are speaking to the customer, for example:
"नमस्ते, क्या मेरी बात {LOADED_CUSTOMER_NAME} जी से हो रही है?"

* If the person confirms they are {LOADED_CUSTOMER_NAME} ("हाँ", "जी हाँ", "speaking", "yes") → continue normally.
* If the person is NOT {LOADED_CUSTOMER_NAME} — wrong number, no such person, or they say it is not them → WRONG PERSON: share no account details, apologize briefly and end the call.
* If the person knows {LOADED_CUSTOMER_NAME} but is busy / unavailable / asks you to call later → CALLBACK: politely say you will call back and end the call, sharing no account details.

A simple name confirmation is the ONLY check. NEVER ask the customer to verify or
provide date of birth, account number, OTP, address, PAN, last payment, or ANY
other identifying detail — at ANY point in the call, not just the opening. Once
they confirm the name, treat identity as fully done and never raise verification
again.

Once the name is confirmed, the greeting and introduction are COMPLETE: never
again open with "नमस्ते", never re-introduce yourself as Priya, and never ask
"क्या मेरी बात ... जी से हो रही है" again — not even after an interruption or a
pause. Returning to the intro after identity is confirmed is a serious error;
instead, carry on from the current point in the conversation.

"""),
        )
        # Last committed user utterance, for the turn-level duplicate backstop
        # in on_user_turn_completed (complements the stt_node final-dedup).
        self._last_user_key = ""
        self._last_user_at = 0.0

    async def on_user_turn_completed(self, turn_ctx, new_message) -> None:
        """Drop a user turn that exactly repeats the previous one.

        A second safety net behind stt_node: if the STT still surfaces the same
        utterance twice in quick succession, we discard the duplicate turn so it
        never reaches the transcript or triggers a second response.
        """
        text = (new_message.text_content or "").strip()
        key = _dedup_key(text)
        now = time.monotonic()
        if (
            key
            and key == self._last_user_key
            and (now - self._last_user_at) < _STT_DEDUP_WINDOW_S
        ):
            logger.info(">>> DUPLICATE TURN dropped: %r", text)
            raise StopResponse()
        self._last_user_key, self._last_user_at = key, now

    async def stt_node(self, audio, model_settings):
        """Filter the STT stream before it becomes a user turn.

        Drops three kinds of final transcripts so they never reach the
        transcript, the LLM, or preemptive generation:
          * garbage (non-Hindi/English STT misfires),
          * filler-only utterances ("hmm", "उह"),
          * back-to-back duplicate finals (the same line recognized twice).
        Everything else (including interim transcripts) passes through
        untouched, so VAD and turn detection are unaffected.
        """
        base = Agent.default.stt_node(self, audio, model_settings)
        if inspect.isawaitable(base):
            base = await base
        if base is None:
            return

        last = {"key": "", "at": 0.0}
        async for event in base:
            if (
                event.type == stt.SpeechEventType.FINAL_TRANSCRIPT
                and event.alternatives
            ):
                text = event.alternatives[0].text or ""

                if not _is_valid_hindi_text(text):
                    logger.error(">>> GARBAGE STT dropped: %r", text)
                    continue
                if _is_filler_only(text):
                    logger.info(">>> FILLER STT dropped: %r", text)
                    continue

                key = _dedup_key(text)
                now = time.monotonic()
                if (
                    key
                    and key == last["key"]
                    and (now - last["at"]) < _STT_DEDUP_WINDOW_S
                ):
                    logger.info(">>> DUPLICATE STT dropped: %r", text)
                    continue
                last["key"], last["at"] = key, now

            yield event

    @function_tool
    async def record_promise_to_pay(
        self,
        context: RunContext,
        amount: str,
        date: str,
    ) -> str:
        """
        Save a confirmed Promise To Pay.

        Call this only after the customer confirms both a payment amount
        and payment date.

        Pass STRUCTURED ENGLISH values, never Hindi text:
          - amount: digits only, in rupees, e.g. "5000" (not "पाँच हज़ार रुपये").
          - date: DD-MM-YYYY, e.g. "25-06-2026". Resolve relative dates such as
            "next Monday" to an absolute date with the year before calling.

        Do not invent values.
        """
        logger.info(
            ">>> PTP-TRACE step=1 tool_dispatched amount=%r date=%r loaded_customer=%r",
            amount,
            date,
            LOADED_CUSTOMER_NAME,
        )

        if not amount.strip() or not date.strip():
            logger.warning(
                "record_promise_to_pay called with missing amount/date; ignoring"
            )
            return "Promise to pay not saved."

        # Normalize defensively so Hindi text can never reach Excel, even if the
        # model passes spoken-form values.
        clean_amount = normalize_ptp_amount(amount)
        clean_date = normalize_ptp_date(date)
        if clean_amount is None or clean_date is None:
            logger.warning(
                "Could not normalize PTP (amount=%r -> %r, date=%r -> %r)",
                amount,
                clean_amount,
                date,
                clean_date,
            )
            return "Promise to pay not saved: amount/date not understood."

        # Speak a short acknowledgment the instant the tool runs, so the customer
        # hears a response instead of silence while the model's spoken
        # confirmation (a second LLM round-trip) is still being generated. Placed
        # after validation so we never acknowledge an unusable promise.
        context.session.say("जी, ठीक है।")

        # Run the blocking Excel write off the event loop so audio is never
        # stalled by file I/O.
        saved = await asyncio.to_thread(
            save_promise_to_pay,
            LOADED_CUSTOMER_NAME,
            clean_amount,
            clean_date,
        )

        if saved:
            return f"PTP saved: {clean_amount} on {clean_date}."

        return "PTP save failed."


server = AgentServer()


def prewarm(proc: JobProcess):
    proc.userdata["vad"] = silero.VAD.load(
        min_speech_duration=0.045,  # 100
        prefix_padding_duration=0.6,  # 200 ms speech pad
        min_silence_duration=0.65,
        activation_threshold=0.26,  # 200 ms silence pad
    )


server.setup_fnc = prewarm


def _first_speech_instructions(name: str) -> str:
    """The opening-turn instruction. नमस्ते is mandated as the FIRST word so the
    call always opens politely (the model used to lead with the intro and skip
    it). Kept as a function so it stays testable."""
    return textwrap.dedent(f"""
Your VERY FIRST word MUST be "नमस्ते" — always open the call politely with it.

Follow this order and do NOT reorder it:
1. Say "नमस्ते".
2. Introduce yourself as Priya from the आईलीड्स financial services company.
3. THEN ask, by name, if you are speaking to {name}.

For example: "नमस्ते, मैं Priya बोल रही हूँ आईलीड्स financial services की तरफ़ से।
क्या मेरी बात {name} जी से हो रही है?"

Speak ONLY in Hindi/Hinglish — never English. Keep it short, natural, and
conversational: one question only.

Do NOT ask for any verification — no date of birth, account number, OTP, address,
or anything else. A simple name confirmation is all you need.
Do NOT congratulate or use words like "बधाई" / "मुबारक" / "congratulations".
Do NOT repeat the introduction again after this.
""")


@server.rtc_session(agent_name=AGENT_NAME)
async def my_agent(ctx: JobContext):

    ctx.log_context_fields = {"room": ctx.room.name}

    session = AgentSession(
        stt=sarvam.STT(
            model="saaras:v3",
            language="hi-IN",
            mode="transcribe",
            high_vad_sensitivity=True,
        ),
        tts=sarvam.TTS(
            target_language_code="hi-IN",
            model="bulbul:v3",
            speaker="ishita",
            pace=1.1,
            # Latency: emit the first audio chunk sooner (default 50, min allowed
            # 30) to cut TTS ttfb, and cache repeated lines (greeting,
            # confirmations) so they come back near-instantly.
            min_buffer_size=30,
            enable_cached_responses=True,
        ),
        turn_detection=TurnDetector(version="v1-mini"),
        vad=ctx.proc.userdata["vad"],
        preemptive_generation=True,
        # Endpointing (end-of-utterance) cap. Default is min_delay=0.5,
        # max_delay=3.0. The Hindi turn detector consistently scores turn-ends
        # below its "user is done" threshold, so every turn was riding up to
        # ~max_delay (measured eou ~2500ms) before the LLM could even start.
        # Capping max_delay to 1.5s removes ~1s of dead air per turn. Tradeoff:
        # a customer who pauses mid-sentence for >1.5s may be cut off — raise
        # this if that happens too often on live calls.
        turn_handling={"endpointing": {"min_delay": 0.3, "max_delay": 0.8}},
    )

    # NOTE: garbage/filler/duplicate STT filtering now lives in
    # Assistant.stt_node (see that method). The previous @session.on("user_speech")
    # handler was a no-op — LiveKit has no "user_speech" event — so it never
    # actually blocked anything; stt_node is the layer that does.

    # ----------------------------------------------------------------------
    # Phase 0 latency instrumentation (baseline measurement).
    #
    # Logs the per-turn breakdown so we can see which stage dominates before
    # tuning anything. The user-perceived "dead air" before Priya replies is
    # roughly:  end_of_utterance_delay + LLM time-to-first-token + TTS
    # time-to-first-byte. EOU/LLM/TTS metrics share a speech_id, so we join them
    # into one line per turn. prompt_cached_tokens shows whether OpenAI prompt
    # caching is hitting (relevant to the Phase 1 prompt-trimming lever).
    # This is read-only logging — remove or gate behind an env var once tuned.
    # ----------------------------------------------------------------------
    turn_latency: dict[str, dict[str, float]] = {}

    @session.on("metrics_collected")
    def _on_metrics(ev: MetricsCollectedEvent) -> None:
        m = ev.metrics
        metrics.log_metrics(m)  # built-in detailed per-metric log

        kind = type(m).__name__
        if kind == "STTMetrics":
            logger.info(
                ">>> LAT stt: duration=%.0fms audio=%.0fms streamed=%s",
                m.duration * 1000,
                m.audio_duration * 1000,
                m.streamed,
            )
            return

        speech_id = getattr(m, "speech_id", None)
        if not speech_id:
            return
        entry = turn_latency.setdefault(speech_id, {})
        if kind == "EOUMetrics":
            entry["eou"] = m.end_of_utterance_delay
        elif kind == "LLMMetrics":
            entry["ttft"] = m.ttft
            entry["prompt_tokens"] = m.prompt_tokens
            entry["cached_tokens"] = m.prompt_cached_tokens
        elif kind == "TTSMetrics":
            entry["ttfb"] = m.ttfb

        if all(k in entry for k in ("eou", "ttft", "ttfb")):
            total = entry["eou"] + entry["ttft"] + entry["ttfb"]
            logger.info(
                ">>> LAT turn speech=%s eou=%.0fms llm_ttft=%.0fms tts_ttfb=%.0fms "
                "~total=%.0fms (prompt=%d tok, cached=%d)",
                speech_id,
                entry["eou"] * 1000,
                entry["ttft"] * 1000,
                entry["ttfb"] * 1000,
                total * 1000,
                int(entry.get("prompt_tokens", 0)),
                int(entry.get("cached_tokens", 0)),
            )
            turn_latency.pop(speech_id, None)

    # End the call once Priya delivers her closing farewell. We detect the
    # farewell in her transcript (deterministically) rather than via a tool call:
    # asking the speaking model to emit a tool call on its closing turn makes the
    # raw call syntax leak into the spoken transcript (and TTS reads it aloud).
    # We wait for the farewell to finish playing before deleting the room.
    farewell_state = {"pending": False, "done": False}
    # Keep strong references to fire-and-forget tasks so they aren't garbage
    # collected mid-flight (see RUF006).
    farewell_tasks: set[asyncio.Task] = set()

    def _spawn(coro):
        task = asyncio.create_task(coro)
        farewell_tasks.add(task)
        task.add_done_callback(farewell_tasks.discard)
        return task

    async def _do_hangup():
        if farewell_state["done"]:
            return
        farewell_state["done"] = True
        # Deleting the room ends both web and SIP sessions and triggers the
        # "close" handler that classifies and saves the outcome.
        await _hangup()

    async def _end_after_farewell():
        # Wait for the farewell utterance to finish playing, then drop the call.
        speech = session.current_speech
        if speech is not None:
            try:
                await speech.wait_for_playout()
            except Exception:
                logger.exception("wait_for_playout failed during farewell hangup")
            await _do_hangup()
        # If the speech handle isn't up yet, the agent_state_changed backstop
        # below fires the hangup once Priya finishes speaking (returns to idle).

    @session.on("conversation_item_added")
    def _detect_farewell(ev):
        item = ev.item
        if getattr(item, "role", None) != "assistant" or farewell_state["pending"]:
            return
        if _is_farewell(item.text_content or ""):
            logger.info(">>> FAREWELL detected; ending call after playout")
            farewell_state["pending"] = True
            _spawn(_end_after_farewell())

    @session.on("agent_state_changed")
    def _hangup_when_done_speaking(ev):
        # Backstop: once the farewell has been flagged and Priya stops speaking,
        # end the call. Guarded by _do_hangup so it never double-fires.
        if farewell_state["pending"] and ev.new_state in ("listening", "idle"):
            _spawn(_do_hangup())

    # Record the call outcome off the voice path, so no tool-call text can leak
    # to the customer. We kick it off on the session "close" event (fires ~2s
    # after hangup) instead of waiting for job shutdown, which LiveKit delays
    # ~20s for room linger. The shutdown callback is only a safety net that
    # awaits the already-running task (no double write, guaranteed completion).
    outcome_state: dict[str, asyncio.Task | None] = {"task": None}

    @session.on("close")
    def _on_session_close(_ev):
        if outcome_state["task"] is None:
            outcome_state["task"] = asyncio.create_task(
                classify_and_save_outcome(session.history, LOADED_CUSTOMER_NAME)
            )

    async def _ensure_outcome_saved():
        task = outcome_state["task"]
        if task is None:
            # close never fired (unusual teardown) — classify now.
            await classify_and_save_outcome(session.history, LOADED_CUSTOMER_NAME)
        elif not task.done():
            await task

    ctx.add_shutdown_callback(_ensure_outcome_saved)

    await session.start(
        agent=Assistant(),
        room=ctx.room,
        # record=True streams audio, transcripts, per-turn traces and logs to
        # LiveKit Agent Observability (Insights tab). Requires the "Agent
        # observability" toggle enabled in the project's Data & privacy settings.
        record=True,
        room_options=room_io.RoomOptions(
            audio_input=room_io.AudioInputOptions(
                noise_cancellation=ai_coustics.audio_enhancement(
                    model=ai_coustics.EnhancerModel.QUAIL_VF_S
                ),
            ),
        ),
    )

    await ctx.connect()

    # Wait until the person is actually in the room before greeting. For the web
    # app they're already present, so this returns immediately; for an outbound
    # phone call it blocks until the callee answers (the SIP participant joins),
    # so Priya never greets an empty room and the customer never picks up to
    # silence.
    await ctx.wait_for_participant()

    # FIRST SPEECH (STRICT CONTROL)
    # allow_interruptions=False: the opening greeting must play fully. On a phone
    # call, line noise / a cough was tripping the turn detector and cutting Priya
    # off mid-introduction, so the greeting is made uninterruptible.
    await session.generate_reply(
        instructions=_first_speech_instructions(LOADED_CUSTOMER_NAME),
        allow_interruptions=False,
    )


if __name__ == "__main__":
    cli.run_app(server)
