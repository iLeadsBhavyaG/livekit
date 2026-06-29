"""Place a test outbound call through the LiveKit -> Plivo trunk.

Flow:
  1. dispatch the `my-agent` worker into a fresh room, and
  2. dial the customer through the Plivo outbound trunk into that same room.

The agent (which must already be running) joins the room and speaks first.

Destination number: defaults to the Phone cell of the customer the agent loads
(the 2nd data row in Customers.xlsx). Override with --to.

Usage (agent must be running in another terminal):
    uv run python src/agent.py dev          # terminal 1
    uv run python telephony/dial.py         # terminal 2  (dials the Excel number)
    uv run python telephony/dial.py --to +9198XXXXXXXX
"""

import argparse
import asyncio
import os
import time
from pathlib import Path

from dotenv import load_dotenv
from livekit import api
from openpyxl import load_workbook

load_dotenv(".env.local")

AGENT_NAME = "my-agent"
DATA_FILE = Path(__file__).parent.parent / "data" / "Customers.xlsx"


def loaded_customer() -> tuple[str | None, str | None]:
    """Return (name, phone) for the row the agent loads (2nd data row)."""
    workbook = load_workbook(DATA_FILE, read_only=True, data_only=True)
    worksheet = workbook.active
    columns: dict[str, int] = {}
    header_found = False
    seen = 0
    name = phone = None
    for row in worksheet.iter_rows(values_only=True):
        if not header_found:
            labels = [str(v).strip().lower() if v is not None else "" for v in row]
            if "customer name" in labels:
                header_found = True
                columns = {lbl: i for i, lbl in enumerate(labels) if lbl}
            continue
        seen += 1
        if seen == 2:  # matches _load_customer_context in src/agent.py
            if "customer name" in columns:
                name = row[columns["customer name"]]
            if "phone" in columns:
                phone = row[columns["phone"]]
            break
    workbook.close()
    return (
        str(name).strip() if name is not None else None,
        str(phone).strip() if phone is not None else None,
    )


async def place_call(to: str, display_name: str | None) -> None:
    trunk_id = os.environ.get("LIVEKIT_SIP_OUTBOUND_TRUNK_ID")
    caller_id = os.environ.get("PLIVO_NUMBER")
    if not trunk_id:
        raise SystemExit(
            "LIVEKIT_SIP_OUTBOUND_TRUNK_ID is not set. "
            "Run telephony/create_outbound_trunk.py first."
        )

    room = f"outbound-{int(time.time())}"
    lk = api.LiveKitAPI()
    try:
        # Named agents only join rooms they are explicitly dispatched to.
        await lk.agent_dispatch.create_dispatch(
            api.CreateAgentDispatchRequest(
                agent_name=AGENT_NAME, room=room, metadata=to
            )
        )
        print(f"Dispatched {AGENT_NAME} -> room {room}; dialing {to} ...")

        # Dial the customer. wait_until_answered surfaces SIP failures (e.g. a
        # trial restriction on an unverified number) as an exception here.
        participant = await lk.sip.create_sip_participant(
            api.CreateSIPParticipantRequest(
                sip_trunk_id=trunk_id,
                sip_call_to=to,
                sip_number=caller_id,  # caller id (the rented Plivo number)
                room_name=room,
                participant_identity="customer",
                participant_name=display_name or "Customer",
                play_ringtone=True,
                wait_until_answered=True,
                krisp_enabled=True,
            )
        )
        print(f"Answered. room={room} participant={participant.participant_identity}")
    finally:
        await lk.aclose()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Outbound test dialer (Plivo).")
    parser.add_argument(
        "--to",
        help="Destination E.164 number; defaults to the loaded customer's Phone cell.",
    )
    args = parser.parse_args()

    name, phone = loaded_customer()
    destination = args.to or phone
    if not destination or "X" in destination:
        raise SystemExit(
            "No valid destination number. Put your (sandboxed) mobile in the loaded "
            "customer's Phone cell in data/Customers.xlsx, or pass --to +9198XXXXXXXX."
        )
    asyncio.run(place_call(destination, name))
